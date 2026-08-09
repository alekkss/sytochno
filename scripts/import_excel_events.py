"""Импорт событий из Excel-файлов comparison_report_*.xlsx в PostgreSQL.

Скрипт находит файлы за последние N дней (по умолчанию 7),
парсит каждый и записывает события в таблицу comparison_events.
Безопасен для повторного запуска — дубли игнорируются через
ON CONFLICT DO NOTHING.

Запуск:
    cd /root/sutochno
    python -m scripts.import_excel_events

Опционально можно передать количество дней через переменную окружения:
    IMPORT_DAYS=14 python -m scripts.import_excel_events
"""

import os
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

# Добавляем корень проекта в sys.path для корректного импорта src.*
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_PROJECT_ROOT))

from src.config.settings import Settings
from src.models.booking_event import (
    BookingEvent,
    CancellationEvent,
    EventType,
    AnyEvent,
)
from src.repositories.postgres_comparison_events_repository import (
    PostgreSQLComparisonEventsRepository,
)


# ── Константы ──

# Количество дней по умолчанию (можно переопределить через IMPORT_DAYS)
_DEFAULT_DAYS: int = 7

# Ожидаемые заголовки Excel (из comparison_export_service.py)
_EXPECTED_HEADERS: list[str] = [
    "Тип события",
    "ID объявления",
    "Название",
    "Дата сделки",
    "Дата заезда",
    "Дата выезда",
    "Ночей",
    "Глубина (дней)",
    "Цена за ночь (руб.)",
    "Итого (руб.)",
]

# Маппинг типов событий из Excel (русский текст) в Enum
_EVENT_TYPE_MAP: dict[str, EventType] = {
    "бронь": EventType.BOOKING,
    "отмена": EventType.CANCELLATION,
}


def _find_excel_files(data_dir: Path, days: int) -> list[Path]:
    """Находит файлы comparison_report_*.xlsx за последние N дней.

    Фильтрует по дате модификации файла (mtime), а не по имени —
    это надёжнее, т.к. формат имени может меняться.

    Args:
        data_dir: Папка с файлами отчётов.
        days: Количество дней назад от текущего момента.

    Returns:
        Список путей к найденным файлам, отсортированный по дате (старые первыми).
    """
    if not data_dir.exists():
        print(f"[ОШИБКА] Папка не найдена: {data_dir}")
        return []

    cutoff = datetime.now(tz=timezone.utc) - timedelta(days=days)
    files: list[tuple[float, Path]] = []

    for path in data_dir.glob("comparison_report_*.xlsx"):
        mtime = path.stat().st_mtime
        file_dt = datetime.fromtimestamp(mtime, tz=timezone.utc)
        if file_dt >= cutoff:
            files.append((mtime, path))

    # Сортировка: старые файлы первыми (хронологический порядок)
    files.sort(key=lambda x: x[0])

    return [path for _, path in files]


def _parse_date_dmy_hm(value: str) -> datetime:
    """Парсит дату в формате 'ДД.ММ.ГГГГ ЧЧ:ММ' → datetime (UTC).

    Args:
        value: Строка с датой и временем.

    Returns:
        datetime с tzinfo=UTC.

    Raises:
        ValueError: Если формат не соответствует ожидаемому.
    """
    dt = datetime.strptime(value.strip(), "%d.%m.%Y %H:%M")
    # Парсер сохраняет snapshot_dt в UTC — считаем, что в Excel тоже UTC
    return dt.replace(tzinfo=timezone.utc)


def _parse_date_dmy(value: str) -> datetime:
    """Парсит дату в формате 'ДД.ММ.ГГГГ' → date.

    Args:
        value: Строка с датой.

    Returns:
        Объект date.

    Raises:
        ValueError: Если формат не соответствует ожидаемому.
    """
    dt = datetime.strptime(value.strip(), "%d.%m.%Y")
    return dt.date()


def _validate_headers(row: tuple) -> bool:
    """Проверяет, что заголовки Excel совпадают с ожидаемыми.

    Args:
        row: Кортеж значений первой строки.

    Returns:
        True если заголовки совпадают.
    """
    actual = [str(cell).strip() if cell else "" for cell in row[:len(_EXPECTED_HEADERS)]]
    return actual == _EXPECTED_HEADERS


def _parse_row(row: tuple, row_idx: int, file_path: Path) -> AnyEvent | None:
    """Парсит одну строку Excel в событие.

    Args:
        row: Кортеж значений строки.
        row_idx: Номер строки (для логирования ошибок).
        file_path: Путь к файлу (для логирования ошибок).

    Returns:
        BookingEvent или CancellationEvent, или None при ошибке парсинга.
    """
    try:
        # Извлекаем значения ячеек
        event_type_raw = str(row[0]).strip().lower() if row[0] else ""
        external_id = str(row[1]).strip() if row[1] else ""
        title = str(row[2]).strip() if row[2] else ""
        deal_dt_raw = str(row[3]).strip() if row[3] else ""
        checkin_raw = str(row[4]).strip() if row[4] else ""
        checkout_raw = str(row[5]).strip() if row[5] else ""
        nights = int(row[6]) if row[6] is not None else 0
        depth_days = int(row[7]) if row[7] is not None else 0
        price_per_night = float(row[8]) if row[8] is not None else 0.0
        total_price = float(row[9]) if row[9] is not None else 0.0

        # Валидация типа события
        if event_type_raw not in _EVENT_TYPE_MAP:
            print(
                f"  [ПРОПУСК] Строка {row_idx}: неизвестный тип события '{event_type_raw}'"
            )
            return None

        event_type = _EVENT_TYPE_MAP[event_type_raw]

        # Валидация обязательных полей
        if not external_id or not deal_dt_raw or not checkin_raw or not checkout_raw:
            print(
                f"  [ПРОПУСК] Строка {row_idx}: пустые обязательные поля"
            )
            return None

        # Парсинг дат
        snapshot_dt = _parse_date_dmy_hm(deal_dt_raw)
        checkin_date = _parse_date_dmy(checkin_raw)
        checkout_date = _parse_date_dmy(checkout_raw)

        # Создание события нужного типа
        if event_type == EventType.BOOKING:
            return BookingEvent(
                listing_external_id=external_id,
                listing_title=title,
                event_type=event_type,
                snapshot_dt=snapshot_dt,
                checkin_date=checkin_date,
                checkout_date=checkout_date,
                nights=nights,
                depth_days=depth_days,
                price_per_night=price_per_night,
                total_price=total_price,
            )
        else:
            return CancellationEvent(
                listing_external_id=external_id,
                listing_title=title,
                event_type=event_type,
                snapshot_dt=snapshot_dt,
                checkin_date=checkin_date,
                checkout_date=checkout_date,
                nights=nights,
                depth_days=depth_days,
                price_per_night=price_per_night,
                total_price=total_price,
            )

    except (ValueError, TypeError, IndexError) as e:
        print(
            f"  [ОШИБКА] Строка {row_idx} в {file_path.name}: {e}"
        )
        return None


def _parse_excel_file(file_path: Path) -> list[AnyEvent]:
    """Парсит один Excel-файл и возвращает список событий.

    Args:
        file_path: Путь к файлу comparison_report_*.xlsx.

    Returns:
        Список событий (может быть пустым при ошибках).
    """
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [ОШИБКА] Не удалось открыть {file_path.name}: {e}")
        return []

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        print(f"  [ПРОПУСК] Файл пуст: {file_path.name}")
        return []

    # Проверка заголовков
    if not _validate_headers(rows[0]):
        print(
            f"  [ОШИБКА] Заголовки не совпадают в {file_path.name}. "
            f"Ожидаемые: {_EXPECTED_HEADERS[:3]}... "
            f"Полученные: {[str(c) for c in rows[0][:3]]}..."
        )
        return []

    # Парсинг строк данных (пропускаем заголовок)
    events: list[AnyEvent] = []
    skipped = 0

    for row_idx, row in enumerate(rows[1:], start=2):
        # Пропускаем пустые строки
        if not row or all(cell is None for cell in row):
            continue

        event = _parse_row(row, row_idx, file_path)
        if event is not None:
            events.append(event)
        else:
            skipped += 1

    if skipped > 0:
        print(f"  Пропущено строк с ошибками: {skipped}")

    return events


def main() -> None:
    """Основная функция импорта событий из Excel в PostgreSQL."""
    print("=" * 60)
    print("Импорт событий из Excel в PostgreSQL")
    print("=" * 60)

    # ── Загрузка конфигурации ──
    try:
        settings = Settings.load()
    except RuntimeError as e:
        print(f"[ОШИБКА] Не удалось загрузить настройки: {e}")
        sys.exit(1)

    if settings.db_type != "postgresql":
        print(
            "[ОШИБКА] Скрипт работает только с DB_TYPE=postgresql. "
            f"Текущий DB_TYPE: '{settings.db_type}'"
        )
        sys.exit(1)

    # ── Определение параметров ──
    days = int(os.getenv("IMPORT_DAYS", str(_DEFAULT_DAYS)))
    data_dir = Path(settings.export_path).parent  # data/

    print(f"\nПапка данных: {data_dir}")
    print(f"Период: последние {days} дней")
    print(f"БД: {settings.pg_host}:{settings.pg_port}/{settings.pg_name}")

    # ── Поиск файлов ──
    files = _find_excel_files(data_dir, days)

    if not files:
        print("\n[INFO] Файлы comparison_report_*.xlsx за указанный период не найдены.")
        print("Нечего импортировать.")
        sys.exit(0)

    print(f"\nНайдено файлов: {len(files)}")
    for f in files:
        mtime = datetime.fromtimestamp(f.stat().st_mtime, tz=timezone.utc)
        print(f"  • {f.name} ({mtime.strftime('%d.%m.%Y %H:%M')} UTC)")

    # ── Парсинг всех файлов ──
    all_events: list[AnyEvent] = []

    print("\n── Парсинг файлов ──")
    for file_path in files:
        print(f"\n  Файл: {file_path.name}")
        events = _parse_excel_file(file_path)
        print(f"  Событий извлечено: {len(events)}")
        all_events.extend(events)

    if not all_events:
        print("\n[INFO] Ни одного события не извлечено из файлов.")
        print("Нечего импортировать.")
        sys.exit(0)

    print(f"\n── Итого событий для импорта: {len(all_events)} ──")

    # Статистика по типам
    bookings = sum(1 for e in all_events if e.event_type == EventType.BOOKING)
    cancellations = len(all_events) - bookings
    print(f"  Броней: {bookings}")
    print(f"  Отмен: {cancellations}")

    # ── Запись в БД ──
    print("\n── Запись в PostgreSQL ──")

    repository = PostgreSQLComparisonEventsRepository(
        dsn=settings.pg_dsn,
        min_pool_size=1,
        max_pool_size=3,
    )

    try:
        repository.initialize()
        print("  Подключение к БД: ОК")

        inserted = repository.bulk_insert(all_events)
        duplicates = len(all_events) - inserted

        print(f"\n  Записано в БД: {inserted}")
        print(f"  Дубликатов (пропущено): {duplicates}")

    except Exception as e:
        print(f"\n[ОШИБКА] Сбой записи в БД: {e}")
        sys.exit(1)
    finally:
        repository.close()
        print("  Пул соединений закрыт.")

    # ── Итоги ──
    print("\n" + "=" * 60)
    print("Импорт завершён успешно!")
    print(f"  Файлов обработано: {len(files)}")
    print(f"  Событий извлечено: {len(all_events)}")
    print(f"  Записано в БД: {inserted}")
    print(f"  Дубликатов: {duplicates}")
    print("=" * 60)


if __name__ == "__main__":
    main()
