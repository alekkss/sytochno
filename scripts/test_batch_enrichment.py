"""Тестовый скрипт batch-обогащения — изолированная проверка этапа 2a.

Читает ID объявлений из существующего Excel-отчёта, запускает
параллельное batch-обогащение через прокси-браузеры и выводит
подробную статистику: время, unbusy/busy/fatal, скорость на объект.

Не использует: репозитории, снимки, сравнение, экспорт, fallback.

Запуск:
    python -m scripts.test_batch_enrichment

Логи пишутся в logs/test_batch_enrichment.log (отдельный файл).
"""

import asyncio
import sys
import time
from pathlib import Path

from openpyxl import load_workbook

from src.config.logger import configure as configure_logging
from src.config.logger import get_logger
from src.config.settings import Settings
from src.models.listing import RawListing
from src.models.proxy import ProxyConfig
from src.services.listing.batch_enrichment_service import BatchEnrichmentService
from src.services.listing.constants import DAYS_COUNT, format_duration
from src.services.proxy_service import ProxyService


# ── Константы скрипта ──────────────────────────────────────

# Путь к Excel-файлу с ID объявлений
_EXCEL_PATH: str = "data/sutochno_report.xlsx"

# Название столбца с ID объявлений в Excel
_ID_COLUMN_NAME: str = "ID объявления"

# Путь к файлу логов этого скрипта
_LOG_FILE_PATH: str = "logs/test_batch_enrichment.log"


def _load_ids_from_excel(excel_path: str) -> list[str]:
    """Загружает ID объявлений из Excel-файла.

    Ищет столбец по заголовку в первой строке, затем читает
    все непустые значения из этого столбца.

    Args:
        excel_path: Путь к Excel-файлу.

    Returns:
        Список ID объявлений (строки).

    Raises:
        FileNotFoundError: Если файл не найден.
        RuntimeError: Если столбец с ID не найден.
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(
            f"Excel-файл не найден: {excel_path}. "
            f"Запустите основную программу хотя бы один раз "
            f"для создания отчёта."
        )

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    if ws is None:
        wb.close()
        raise RuntimeError(f"В файле {excel_path} нет активного листа.")

    # Ищем столбец с ID по заголовку в первой строке
    id_col_idx: int | None = None
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))

    for idx, cell in enumerate(header_row):
        if cell.value is not None and str(cell.value).strip() == _ID_COLUMN_NAME:
            id_col_idx = idx
            break

    if id_col_idx is None:
        wb.close()
        raise RuntimeError(
            f"Столбец '{_ID_COLUMN_NAME}' не найден в файле {excel_path}. "
            f"Найденные заголовки: "
            f"{[str(c.value).strip() for c in header_row if c.value is not None]}"
        )

    # Читаем все ID начиная со второй строки
    ids: list[str] = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        cell = row[id_col_idx]
        if cell.value is not None:
            raw_value = str(cell.value).strip()
            # Убираем дробную часть если число пришло как float
            if "." in raw_value:
                try:
                    raw_value = str(int(float(raw_value)))
                except (ValueError, TypeError):
                    pass
            if raw_value:
                ids.append(raw_value)

    wb.close()
    return ids


def _create_listings_from_ids(ids: list[str]) -> list[RawListing]:
    """Создаёт минимальные RawListing объекты из списка ID.

    Заполняет только обязательные поля (external_id, title, url) —
    этого достаточно для batch-обогащения, которое работает
    исключительно через API по ID объявлений.

    Args:
        ids: Список ID объявлений.

    Returns:
        Список RawListing с минимальными данными.
    """
    listings: list[RawListing] = []

    for ext_id in ids:
        listing = RawListing(
            external_id=ext_id,
            title=f"Объявление {ext_id}",
            url=f"https://sutochno.ru/{ext_id}",
        )
        listings.append(listing)

    return listings


def _print_statistics(
    listings: list[RawListing],
    elapsed: float,
    proxy_count: int,
) -> None:
    """Выводит подробную статистику batch-обогащения в консоль.

    Args:
        listings: Список карточек после обогащения.
        elapsed: Общее время выполнения (секунды).
        proxy_count: Количество использованных прокси-воркеров.
    """
    total = len(listings)

    # ── Классификация результатов ──
    unbusy_count = 0      # Полностью свободные (календарь = все 0, есть цены)
    busy_count = 0        # Частично занятые (есть 1 в календаре, есть цены)
    fatal_count = 0       # Фатальные ошибки (enrichment_skip_reason)
    unenriched_count = 0  # Не обогащённые (нет данных, нет фатальной причины)
    prices_only_count = 0  # Есть цены, но нет календаря

    # Подсчёт по причинам фатальных ошибок
    fatal_reasons: dict[str, int] = {}

    for listing in listings:
        if listing.enrichment_skip_reason is not None:
            fatal_count += 1
            reason = listing.enrichment_skip_reason
            fatal_reasons[reason] = fatal_reasons.get(reason, 0) + 1
            continue

        has_calendar = bool(listing.calendar_60_days) and len(listing.calendar_60_days) == DAYS_COUNT
        has_prices = bool(listing.prices_60_days) and any(p > 0 for p in listing.prices_60_days)
        has_busy_days = has_calendar and any(c == 1 for c in listing.calendar_60_days)
        all_free = has_calendar and all(c == 0 for c in listing.calendar_60_days)

        if all_free and has_prices:
            unbusy_count += 1
        elif has_busy_days and has_prices:
            busy_count += 1
        elif has_prices and not has_calendar:
            prices_only_count += 1
        else:
            unenriched_count += 1

    enriched_total = unbusy_count + busy_count + prices_only_count

    # ── Вывод ──
    separator = "═" * 60

    print(f"\n{separator}")  # noqa: T201
    print("  РЕЗУЛЬТАТЫ BATCH-ОБОГАЩЕНИЯ (ЭТАП 2a)")  # noqa: T201
    print(separator)  # noqa: T201

    print(f"\n  Всего объявлений:        {total}")  # noqa: T201
    print(f"  Прокси-воркеров:         {proxy_count}")  # noqa: T201
    print(f"  Общее время:             {format_duration(elapsed)}")  # noqa: T201

    if total > 0:
        speed = elapsed / total
        print(f"  Среднее время на объект: {speed:.2f}с")  # noqa: T201
        throughput = total / elapsed * 60 if elapsed > 0 else 0
        print(f"  Пропускная способность:  {throughput:.0f} объектов/мин")  # noqa: T201

    print(f"\n{'─' * 60}")  # noqa: T201
    print("  КЛАССИФИКАЦИЯ РЕЗУЛЬТАТОВ")  # noqa: T201
    print(f"{'─' * 60}")  # noqa: T201

    print(f"\n  Обогащено всего:         {enriched_total} ({_pct(enriched_total, total)})")  # noqa: T201
    print(f"    ├─ unbusy (свободные): {unbusy_count} ({_pct(unbusy_count, total)})")  # noqa: T201
    print(f"    ├─ busy (занятые):     {busy_count} ({_pct(busy_count, total)})")  # noqa: T201
    print(f"    └─ только цены:        {prices_only_count} ({_pct(prices_only_count, total)})")  # noqa: T201

    print(f"\n  Фатальные ошибки:        {fatal_count} ({_pct(fatal_count, total)})")  # noqa: T201
    if fatal_reasons:
        for reason, count in sorted(fatal_reasons.items(), key=lambda x: -x[1]):
            print(f"    └─ {reason}: {count}")  # noqa: T201

    print(f"\n  Не обогащены (fallback): {unenriched_count} ({_pct(unenriched_count, total)})")  # noqa: T201

    # ── Оценка необходимости прогрева ──
    print(f"\n{'─' * 60}")  # noqa: T201
    print("  ВЫВОДЫ")  # noqa: T201
    print(f"{'─' * 60}")  # noqa: T201

    if enriched_total == 0 and total > 0:
        print("\n  ⚠ Ни одна карточка не обогащена!")  # noqa: T201
        print("    Возможные причины:")  # noqa: T201
        print("    - Все прокси забанены (токены не получены)")  # noqa: T201
        print("    - API изменил формат ответа")  # noqa: T201
        print("    - Требуется прогрев сессии перед batch-запросами")  # noqa: T201
    elif unenriched_count > total * 0.3:
        print(f"\n  ⚠ Высокий процент необогащённых: {_pct(unenriched_count, total)}")  # noqa: T201
        print("    Возможно, часть прокси забанена или токены протухают.")  # noqa: T201
        print("    Рекомендуется увеличить количество прокси или")  # noqa: T201
        print("    добавить прогрев (задержку после получения токена).")  # noqa: T201
    else:
        print(f"\n  ✓ Batch-обогащение работает стабильно.")  # noqa: T201
        print(f"    Обогащено {_pct(enriched_total, total)} карточек.")  # noqa: T201
        if unenriched_count > 0:
            print(f"    {unenriched_count} карточек для fallback-обогащения.")  # noqa: T201

    # ── Примеры обогащённых карточек ──
    enriched_samples = [
        l for l in listings
        if l.enrichment_skip_reason is None
        and l.calendar_60_days
        and l.prices_60_days
        and any(p > 0 for p in l.prices_60_days)
    ]

    if enriched_samples:
        print(f"\n{'─' * 60}")  # noqa: T201
        print("  ПРИМЕРЫ ОБОГАЩЁННЫХ КАРТОЧЕК (до 5 шт.)")  # noqa: T201
        print(f"{'─' * 60}")  # noqa: T201

        for sample in enriched_samples[:5]:
            occ = sample.occupancy_percent
            avg = sample.average_price
            busy_days = sum(sample.calendar_60_days) if sample.calendar_60_days else 0
            free_days = DAYS_COUNT - busy_days
            cal_preview = "".join(str(c) for c in sample.calendar_60_days[:30])

            print(f"\n  ID: {sample.external_id}")  # noqa: T201
            print(f"    Занятость:     {occ}% ({busy_days} занято, {free_days} свободно)")  # noqa: T201
            print(f"    Средняя цена:  {avg} руб./сут.")  # noqa: T201
            print(f"    Календарь:     {cal_preview}... (первые 30 дней)")  # noqa: T201

    print(f"\n{separator}\n")  # noqa: T201


def _pct(value: int, total: int) -> str:
    """Форматирует процент от общего числа.

    Args:
        value: Числитель.
        total: Знаменатель.

    Returns:
        Строка вида «42.5%» или «0.0%».
    """
    if total == 0:
        return "0.0%"
    return f"{(value / total) * 100:.1f}%"


async def run() -> None:
    """Основная асинхронная логика тестового скрипта."""
    # --- Шаг 1: Загрузка конфигурации ---
    try:
        settings = Settings.load()
    except RuntimeError as e:
        print(f"[ОШИБКА] {e}", file=sys.stderr)  # noqa: T201
        sys.exit(1)

    # --- Шаг 2: Конфигурация логирования ---
    configure_logging(
        log_level="DEBUG",
        log_file_path=_LOG_FILE_PATH,
    )
    logger = get_logger("test_batch")

    logger.info(
        "тест_batch_обогащения_запущен",
        step="init",
    )

    # --- Шаг 3: Загрузка ID из Excel ---
    print(f"\n📂 Загрузка ID из {_EXCEL_PATH}...")  # noqa: T201

    try:
        ids = _load_ids_from_excel(_EXCEL_PATH)
    except (FileNotFoundError, RuntimeError) as e:
        print(f"[ОШИБКА] {e}", file=sys.stderr)  # noqa: T201
        sys.exit(1)

    print(f"   Загружено ID: {len(ids)}")  # noqa: T201

    logger.info(
        "id_загружены_из_excel",
        total=len(ids),
        step="init",
    )

    if not ids:
        print("[ОШИБКА] В Excel-файле нет ID объявлений.", file=sys.stderr)  # noqa: T201
        sys.exit(1)

    # --- Шаг 4: Создание RawListing из ID ---
    listings = _create_listings_from_ids(ids)

    print(f"   Создано карточек: {len(listings)}")  # noqa: T201

    # --- Шаг 5: Загрузка и проверка прокси ---
    if not settings.use_proxy:
        print(  # noqa: T201
            "[ОШИБКА] USE_PROXY=false в .env. "
            "Этот скрипт тестирует параллельное batch-обогащение через прокси. "
            "Установите USE_PROXY=true и укажите прокси в файле.",
            file=sys.stderr,
        )
        sys.exit(1)

    print("\n🔍 Проверка прокси...")  # noqa: T201

    proxy_service = ProxyService(settings=settings)

    try:
        proxies = proxy_service.load_proxies()
    except RuntimeError as e:
        print(f"[ОШИБКА] {e}", file=sys.stderr)  # noqa: T201
        sys.exit(1)

    print(f"   Загружено прокси: {len(proxies)}")  # noqa: T201

    working_proxies: list[ProxyConfig] = await proxy_service.check_proxies(proxies)

    print(f"   Рабочих прокси: {len(working_proxies)}")  # noqa: T201

    if not working_proxies:
        print(  # noqa: T201
            "[ОШИБКА] Нет рабочих прокси. Проверьте файл прокси.",
            file=sys.stderr,
        )
        sys.exit(1)

    logger.info(
        "прокси_готовы",
        total=len(working_proxies),
        max_workers=settings.max_proxy_workers,
        step="proxy",
    )

    max_workers = min(len(working_proxies), settings.max_proxy_workers)

    # --- Шаг 6: Запуск batch-обогащения ---
    first_search_url = settings.search_urls[0]

    print(f"\n🚀 Запуск batch-обогащения...")  # noqa: T201
    print(f"   Карточек:   {len(listings)}")  # noqa: T201
    print(f"   Воркеров:   {max_workers}")  # noqa: T201
    print(f"   Поисковый URL: {first_search_url[:80]}...")  # noqa: T201
    print(f"   Ожидайте завершения...\n")  # noqa: T201

    batch_service = BatchEnrichmentService()

    start_time = time.perf_counter()

    try:
        await batch_service.enrich_batch_parallel(
            settings=settings,
            listings=listings,
            proxies=working_proxies,
            search_url=first_search_url,
            proxy_service=proxy_service,
        )
    except Exception as e:
        logger.exception(
            "ошибка_batch_обогащения",
            error=str(e),
            error_type=type(e).__name__,
        )
        print(f"\n[ОШИБКА] Batch-обогащение завершилось с ошибкой: {e}", file=sys.stderr)  # noqa: T201

    elapsed = time.perf_counter() - start_time

    logger.info(
        "тест_batch_завершён",
        elapsed=format_duration(elapsed),
        total=len(listings),
        step="done",
    )

    # --- Шаг 7: Вывод статистики ---
    _print_statistics(listings, elapsed, max_workers)

    print(f"📄 Подробные логи: {_LOG_FILE_PATH}")  # noqa: T201


def main() -> None:
    """Синхронная точка входа."""
    asyncio.run(run())


if __name__ == "__main__":
    main()
