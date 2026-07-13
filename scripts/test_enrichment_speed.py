"""Диагностический скрипт — измерение скорости и стабильности Этапа 2.

Прогрев браузера выполняется через тот же метод, что и в боевом
пайплайне — EnrichStrategies._warmup_browser() (навигация на главную
sutochno.ru + пауза, до 2 попыток с заменой прокси при неудаче).
Метод переиспользуется напрямую, а не дублируется — это гарантирует,
что диагностика измеряет ту же самую «прогретую» сессию, что и боевой
парсинг.

Для каждой карточки из списка ID (после прогрева воркера):
1. Загружает страницу и перехватывает токен — замеряет время.
2. Делает тестовый API-запрос (getPricesAndAvailabilities на 2 ночи) — замеряет время.
3. Логирует причину ошибки при неудаче.

Данные НЕ собираются, снимки НЕ создаются — чистая диагностика.

Запуск:
    python -m scripts.test_enrichment_speed
"""

import asyncio
import os
import statistics
import sys
import time
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path

import openpyxl

# ── Добавляем корень проекта в sys.path ──
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.config.logger import configure, get_logger
from src.config.settings import Settings
from src.models.proxy import ProxyConfig
from src.services.browser_service import BrowserService
from src.services.listing.enrich_strategies import EnrichStrategies
from src.services.proxy_service import ProxyService

logger = get_logger("test_speed")

# ── Источник списка ID карточек для тестирования ──
# ID больше не хардкодятся в скрипте — читаются из Excel-отчёта проекта.
# Путь и название столбца настраиваются через .env, чтобы не привязывать
# скрипт к конкретному расположению файла.
_IDS_SOURCE_PATH_ENV: str = "IDS_SOURCE_XLSX_PATH"
_IDS_SOURCE_COLUMN_ENV: str = "IDS_SOURCE_COLUMN_NAME"

# Дефолтный путь — относительно корня проекта (см. .env.example).
_DEFAULT_IDS_SOURCE_PATH: str = "sutochno_report.xlsx"
_DEFAULT_IDS_SOURCE_COLUMN: str = "ID объявления"

# Максимальное количество карточек для теста (из файла с ID).
# Установите 0 для обработки всех.
MAX_TEST_CARDS: int = 1000

# Переопределение количества параллельных вкладок для диагностики.
# 0 = использовать settings.max_tabs (значение MAX_TABS из .env),
# как в боевом пайплайне.
# По итогам первого прогона: волна из 5 одновременных вкладок на
# свежем прокси почти всегда проваливается целиком (см. data/
# diag_enrichment_speed_with_warmup.csv, воркеры 3-6, волна 0).
# Снижаем до 2, не меняя боевой MAX_TABS в .env.
_DIAG_MAX_TABS_OVERRIDE: int = 2

# URL шаблон для карточки (прямой URL фронтенда)
_ENRICHMENT_URL_TEMPLATE: str = (
    "https://sutochno.ru/front/searchapp/detail/{object_id}"
)

# URL API для тестового запроса
_API_PRICES_URL: str = (
    "https://sutochno.ru/api/json/objects/getPricesAndAvailabilities"
)

# Таймаут ожидания токена после domcontentloaded (секунды)
_TOKEN_WAIT_SECONDS: float = 10.0

# Интервал поллинга токена (секунды)
_TOKEN_POLL_INTERVAL: float = 0.2

# ── Параметры плавности запуска ──

# Максимальное количество одновременно запускаемых браузеров.
_MAX_CONCURRENT_BROWSER_LAUNCHES: int = 3

# Пауза между последовательными запусками браузеров внутри семафора (сек).
_BROWSER_LAUNCH_STAGGER_SECONDS: float = 2.0

# Задержка между стартом вкладок внутри воркера (секунды).
_TAB_START_DELAY_SECONDS: float = 2.0

# Пауза между порциями вкладок (секунды).
_INTER_CHUNK_PAUSE_SECONDS: float = 1.0

# ── Параметры завершения ──

# Максимальное время ожидания завершения всех pending-задач
# после gather при выходе из main() (секунды).
# Playwright внутри создаёт фоновые asyncio.Task для обработки
# CDP-событий — при закрытии 80 браузеров часть этих задач
# остаётся в pending и блокирует asyncio.run(). Этот таймаут
# гарантирует, что скрипт завершится за предсказуемое время.
_SHUTDOWN_PENDING_TIMEOUT: float = 10.0

# Максимальное количество одновременных остановок браузеров.
# Playwright.stop() закрывает pipe к Node.js-драйверу — если
# 80 процессов Node.js закрываются одновременно, ОС может
# исчерпать файловые дескрипторы или замедлить обработку сигналов.
# Ограничение до 5 одновременных остановок предотвращает это.
_MAX_CONCURRENT_STOPS: int = 5


# ── Модель результата диагностики ──


@dataclass
class CardDiagResult:
    """Результат диагностики одной карточки."""

    object_id: str
    worker_idx: int = 0
    is_first_card: bool = False

    # Этап 1: Загрузка страницы
    page_loaded: bool = False
    page_load_time_s: float = 0.0
    page_error: str = ""

    # Этап 2: Перехват токена
    token_captured: bool = False
    token_wait_time_s: float = 0.0

    # Этап 3: API-запрос
    api_success: bool = False
    api_response_time_s: float = 0.0
    api_error: str = ""
    api_busy_status: str = ""

    # Общее время
    total_time_s: float = 0.0


@dataclass
class WorkerWarmupInfo:
    """Результат прогрева браузера воркера (EnrichStrategies._warmup_browser).

    Прогрев выполняется ДО обработки карточек — навигация на главную
    страницу sutochno.ru устанавливает сессионные cookies, без которых
    фронтенд карточки не инициализирует API-запросы.
    """

    worker_idx: int
    proxy_label: str
    success: bool = False
    time_s: float = 0.0
    error: str = ""


@dataclass
class WorkerStats:
    """Статистика одного воркера."""

    worker_idx: int
    proxy: str
    cards_total: int = 0
    cards_page_ok: int = 0
    cards_token_ok: int = 0
    cards_api_ok: int = 0
    errors_by_reason: dict[str, int] = field(default_factory=dict)
    page_load_times: list[float] = field(default_factory=list)
    token_wait_times: list[float] = field(default_factory=list)
    api_response_times: list[float] = field(default_factory=list)


# ── Классификация сетевых ошибок ──

_ERROR_CLASSIFIERS: list[tuple[str, str]] = [
    ("ERR_TUNNEL_CONNECTION_FAILED", "ERR_TUNNEL_CONNECTION_FAILED"),
    ("ERR_PROXY_CONNECTION_FAILED", "ERR_PROXY_CONNECTION_FAILED"),
    ("ERR_TIMED_OUT", "ERR_TIMED_OUT"),
    ("ERR_CONNECTION_RESET", "ERR_CONNECTION_RESET"),
    ("ERR_CONNECTION_REFUSED", "ERR_CONNECTION_REFUSED"),
    ("ERR_CONNECTION_CLOSED", "ERR_CONNECTION_CLOSED"),
    ("ERR_EMPTY_RESPONSE", "ERR_EMPTY_RESPONSE"),
    ("Timeout", "navigation_timeout"),
]


def _classify_page_error(error_msg: str, timeout_ms: int) -> str:
    """Классифицирует ошибку загрузки страницы."""
    for marker, label in _ERROR_CLASSIFIERS:
        if marker in error_msg:
            if label == "navigation_timeout":
                return f"timeout_{timeout_ms}ms"
            return label

    if "net::" in error_msg:
        for part in error_msg.split():
            if part.startswith("net::"):
                return part
        return f"net_error: {error_msg[:80]}"

    return f"{error_msg[:100]}"


# ── Загрузка ID объявлений из Excel-отчёта ──


def load_test_ids_from_xlsx(path: Path, column_name: str) -> list[str]:
    """Загружает список ID объявлений из Excel-отчёта проекта.

    Ищет столбец column_name в первой строке (заголовки) активного
    листа, затем читает все непустые значения ниже него как ID.
    Дубликаты отбрасываются, порядок появления в файле сохраняется.

    Args:
        path: Путь к xlsx-файлу.
        column_name: Точное название столбца с ID объявления.

    Returns:
        Список ID в виде строк.

    Raises:
        RuntimeError: Если файл не найден, не открывается, столбец
            не найден или в столбце нет ни одного ID.
    """
    if not path.exists():
        raise RuntimeError(
            f"Файл с ID объявлений не найден: {path}. "
            f"Проверьте переменную {_IDS_SOURCE_PATH_ENV} в .env."
        )

    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise RuntimeError(
            f"Не удалось открыть xlsx-файл {path}: {type(e).__name__}: {e}"
        ) from e

    try:
        sheet = workbook.active
        header_row = next(
            sheet.iter_rows(min_row=1, max_row=1, values_only=True), None
        )

        if not header_row:
            raise RuntimeError(f"Файл {path} пуст — нет строки заголовков.")

        column_index: int | None = None
        for idx, header in enumerate(header_row):
            if header is not None and str(header).strip() == column_name:
                column_index = idx
                break

        if column_index is None:
            available = [str(h) for h in header_row if h is not None]
            raise RuntimeError(
                f"Столбец '{column_name}' не найден в файле {path}. "
                f"Доступные столбцы: {available}. "
                f"Проверьте переменную {_IDS_SOURCE_COLUMN_ENV} в .env."
            )

        ids: list[str] = []
        seen: set[str] = set()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if column_index >= len(row):
                continue

            raw_value = row[column_index]
            if raw_value is None or str(raw_value).strip() == "":
                continue

            # Excel хранит числовые ID как float (например 2298659.0) —
            # приводим к целому перед строковым представлением, иначе
            # ID перестанет совпадать с тем, что ожидает API sutochno.ru.
            if isinstance(raw_value, float) and raw_value.is_integer():
                object_id = str(int(raw_value))
            else:
                object_id = str(raw_value).strip()

            if object_id and object_id not in seen:
                seen.add(object_id)
                ids.append(object_id)
    finally:
        workbook.close()

    if not ids:
        raise RuntimeError(
            f"В столбце '{column_name}' файла {path} не найдено ни одного ID."
        )

    logger.info(
        "id_загружены_из_xlsx",
        step=f"файл={path}, столбец='{column_name}'",
        total=len(ids),
    )
    return ids


# ── Диагностика одной карточки ──


async def diagnose_card(
    page: "any",  # type: ignore[name-defined]
    object_id: str,
    navigation_timeout_ms: int,
    worker_idx: int = 0,
    is_first_card: bool = False,
) -> CardDiagResult:
    """Диагностирует одну карточку: загрузка, токен, API-запрос."""
    result = CardDiagResult(
        object_id=object_id,
        worker_idx=worker_idx,
        is_first_card=is_first_card,
    )
    card_start = time.perf_counter()

    url = _ENRICHMENT_URL_TEMPLATE.format(object_id=object_id)
    captured_token: list[str] = []

    async def _route_handler(route: "any") -> None:  # type: ignore[name-defined]
        req = route.request
        if "sutochno.ru/api/json" in req.url:
            token = req.headers.get("token") or req.headers.get("Token")
            if token and not captured_token:
                captured_token.append(token)
        try:
            await route.continue_()
        except Exception:
            pass

    try:
        await page.route("**/api/json/**", _route_handler)
    except Exception as e:
        result.page_error = f"route_setup: {type(e).__name__}"
        result.total_time_s = time.perf_counter() - card_start
        return result

    # ── Этап 1: Загрузка страницы ──
    page_start = time.perf_counter()
    try:
        await page.goto(
            url,
            wait_until="domcontentloaded",
            timeout=navigation_timeout_ms,
        )
        result.page_loaded = True
        result.page_load_time_s = time.perf_counter() - page_start

        current_url = page.url
        if "sutochno.ru" not in current_url:
            result.page_loaded = False
            result.page_error = f"redirect: {current_url[:80]}"

    except Exception as e:
        result.page_load_time_s = time.perf_counter() - page_start
        result.page_error = _classify_page_error(str(e), navigation_timeout_ms)

    # ── Этап 2: Ожидание токена ──
    if result.page_loaded:
        token_start = time.perf_counter()

        if not captured_token:
            elapsed = 0.0
            while elapsed < _TOKEN_WAIT_SECONDS:
                if captured_token:
                    break
                await asyncio.sleep(_TOKEN_POLL_INTERVAL)
                elapsed += _TOKEN_POLL_INTERVAL

        result.token_wait_time_s = time.perf_counter() - token_start
        result.token_captured = bool(captured_token)

    try:
        await page.unroute("**/api/json/**")
    except Exception:
        pass

    # ── Этап 3: Тестовый API-запрос ──
    if result.token_captured:
        token = captured_token[0]
        today = date.today()
        date_begin = f"{today} 14:00:00"
        date_end = f"{today + timedelta(days=2)} 11:00:00"

        api_body = {
            "objects": [int(object_id)],
            "rooms_cnt": {},
            "guests": 2,
            "date_begin": date_begin,
            "date_end": date_end,
            "currency_id": 1,
            "is_pets": 0,
            "documents": 0,
            "target": 0,
            "ages": [],
            "no_time": 1,
        }

        api_start = time.perf_counter()

        try:
            api_result = await page.evaluate("""
                async ({url, body, token}) => {
                    try {
                        const resp = await fetch(url, {
                            method: 'POST',
                            headers: {
                                'Content-Type': 'application/json',
                                'Accept': 'application/json',
                                'token': token,
                                'platform': 'js',
                                'api-version': '1.13'
                            },
                            body: JSON.stringify(body),
                            credentials: 'include'
                        });
                        const text = await resp.text();
                        try {
                            const data = JSON.parse(text);
                            return {success: true, status: resp.status, data: data};
                        } catch(e) {
                            return {success: false, error: 'JSON parse error', raw: text.substring(0, 300)};
                        }
                    } catch(e) {
                        return {success: false, error: e.message};
                    }
                }
            """, {"url": _API_PRICES_URL, "body": api_body, "token": token})

            result.api_response_time_s = time.perf_counter() - api_start

            if api_result.get("success"):
                data = api_result.get("data", {})
                if isinstance(data, dict) and data.get("success"):
                    objects = data.get("data", {}).get("objects", [])
                    if objects and isinstance(objects, list):
                        obj_data = objects[0]
                        if isinstance(obj_data, dict):
                            inner = obj_data.get("data", {})
                            if isinstance(inner, dict):
                                result.api_success = True
                                result.api_busy_status = inner.get("busy", "unknown")
                            else:
                                result.api_error = "no_inner_data"
                        else:
                            result.api_error = "invalid_object_format"
                    else:
                        result.api_error = "no_objects"
                else:
                    errors = data.get("errors", []) if isinstance(data, dict) else []
                    result.api_error = f"api_error: {str(errors)[:150]}"
            else:
                result.api_error = api_result.get("error", "unknown_fetch_error")

        except Exception as e:
            result.api_response_time_s = time.perf_counter() - api_start
            result.api_error = f"{type(e).__name__}: {str(e)[:100]}"

    result.total_time_s = time.perf_counter() - card_start

    first_marker = " [ПЕРВАЯ]" if is_first_card else ""
    if result.api_success:
        logger.debug(
            "карточка_ок",
            step=f"id={object_id}, в={worker_idx}{first_marker}, "
                 f"стр={result.page_load_time_s:.1f}с, "
                 f"токен={result.token_wait_time_s:.1f}с, "
                 f"api={result.api_response_time_s:.1f}с, "
                 f"busy={result.api_busy_status}",
        )
    else:
        error = result.page_error or result.api_error or "token_not_captured"
        logger.debug(
            "карточка_ошибка",
            step=f"id={object_id}, в={worker_idx}{first_marker}, причина={error}",
        )

    return result


# ── Глобальный семафор запуска браузеров ──

_launch_semaphore: asyncio.Semaphore | None = None
_launch_lock: asyncio.Lock | None = None
_last_launch_time: float = 0.0


async def _acquire_launch_slot(worker_idx: int) -> None:
    """Ожидает слот для запуска браузера с плавной задержкой."""
    global _last_launch_time  # noqa: PLW0603

    assert _launch_semaphore is not None
    assert _launch_lock is not None

    await _launch_semaphore.acquire()

    async with _launch_lock:
        now = time.monotonic()
        since_last = now - _last_launch_time
        if since_last < _BROWSER_LAUNCH_STAGGER_SECONDS:
            wait = _BROWSER_LAUNCH_STAGGER_SECONDS - since_last
            logger.debug(
                "ожидание_слота_запуска",
                step=f"в={worker_idx}, пауза={wait:.1f}с",
            )
            await asyncio.sleep(wait)
        _last_launch_time = time.monotonic()

    logger.info(
        "слот_запуска_получен",
        step=f"в={worker_idx}",
    )


def _release_launch_slot() -> None:
    """Освобождает слот запуска после старта браузера."""
    assert _launch_semaphore is not None
    _launch_semaphore.release()


# ── Воркер ──


async def run_worker(
    worker_idx: int,
    card_ids: list[str],
    settings: Settings,
    proxy: ProxyConfig | None = None,
    max_tabs: int = 1,
    proxy_service: ProxyService | None = None,
    all_proxies: list[ProxyConfig] | None = None,
) -> tuple[list[CardDiagResult], BrowserService, WorkerWarmupInfo]:
    """Воркер — запускает браузер, прогревает его и идёт на карточки.

    Прогрев выполняется через EnrichStrategies._warmup_browser() —
    тот же метод, что и в боевом пайплайне (навигация на главную
    sutochno.ru + пауза, до 2 попыток с заменой прокси при неудаче).
    Метод переиспользуется напрямую, а не дублируется — это гарантирует,
    что первая карточка воркера обрабатывается в такой же «прогретой»
    сессии, как и в продовом enrich_listings_parallel, а не в холодной.

    ВАЖНО: возвращает BrowserService вместе с результатами.
    Остановка браузера выполняется централизованно в main() —
    это предотвращает гонку exception_handler'ов при параллельном
    закрытии 80 Playwright-инстансов.

    Args:
        worker_idx: Номер воркера.
        card_ids: Список ID карточек.
        settings: Настройки приложения.
        proxy: Прокси (None = без прокси).
        max_tabs: Количество параллельных вкладок.
        proxy_service: Сервис прокси — используется _warmup_browser
            для проверки/замены прокси при неудачном прогреве.
        all_proxies: Полный список прокси всех воркеров — нужен
            _warmup_browser для исключения занятых при поиске замены.

    Returns:
        Кортеж (список результатов диагностики, BrowserService,
        результат прогрева).
    """
    results: list[CardDiagResult] = []
    browser_service = BrowserService(settings=settings)
    nav_timeout = settings.navigation_timeout
    proxy_label = str(proxy) if proxy else "без_прокси"
    warmup_info = WorkerWarmupInfo(worker_idx=worker_idx, proxy_label=proxy_label)
    launch_slot_held = False
    cards_processed = 0

    try:
        # ── Фаза 1: Ожидание слота запуска ──
        await _acquire_launch_slot(worker_idx)
        launch_slot_held = True

        # ── Фаза 2: Запуск браузера ──
        logger.info(
            "воркер_запуск_браузера",
            step=f"в={worker_idx}, прокси={proxy_label}, "
                 f"карточек={len(card_ids)}",
        )

        await browser_service.start(proxy=proxy)

        # Освобождаем слот сразу после запуска браузера —
        # следующий воркер может начинать запуск
        _release_launch_slot()
        launch_slot_held = False

        logger.info(
            "воркер_браузер_готов",
            step=f"в={worker_idx}, прокси={proxy_label}",
        )

        # ── Фаза 2.5: Прогрев браузера (как в боевом пайплайне) ──
        # EnrichStrategies._warmup_browser может заменить прокси на
        # рабочую из пула, если исходная не отвечает — поэтому proxy
        # и proxy_label обновляются по результату вызова.
        warmup_start = time.perf_counter()
        try:
            warmup_ok, active_proxy = await EnrichStrategies._warmup_browser(
                browser_service=browser_service,
                proxy=proxy,
                worker_idx=worker_idx,
                all_proxies=all_proxies or [],
                proxy_service=proxy_service,
            )
        except Exception as e:
            warmup_ok = False
            active_proxy = proxy
            warmup_info.error = f"{type(e).__name__}: {str(e)[:150]}"

        warmup_info.time_s = time.perf_counter() - warmup_start
        warmup_info.success = warmup_ok
        proxy = active_proxy
        proxy_label = str(proxy) if proxy else "без_прокси"
        warmup_info.proxy_label = proxy_label

        if not warmup_ok:
            logger.error(
                "воркер_прогрев_не_удался_обработка_отменена",
                step=f"в={worker_idx}, время={warmup_info.time_s:.1f}с",
            )
            return (results, browser_service, warmup_info)

        logger.info(
            "воркер_прогрет",
            step=f"в={worker_idx}, прокси={proxy_label}, "
                 f"время={warmup_info.time_s:.1f}с",
        )

        # ── Фаза 3: Обработка карточек (первая = после прогрева) ──
        for chunk_start in range(0, len(card_ids), max_tabs):
            chunk = card_ids[chunk_start: chunk_start + max_tabs]

            async def _process_tab(
                idx: int, card_id: str,
            ) -> CardDiagResult:
                page = None
                try:
                    if idx > 0:
                        await asyncio.sleep(_TAB_START_DELAY_SECONDS)
                    page = await browser_service.create_page()

                    # Помечаем первую карточку воркера
                    is_first = (cards_processed + idx == 0)

                    return await diagnose_card(
                        page=page,
                        object_id=card_id,
                        navigation_timeout_ms=nav_timeout,
                        worker_idx=worker_idx,
                        is_first_card=is_first,
                    )
                except Exception as e:
                    return CardDiagResult(
                        object_id=card_id,
                        worker_idx=worker_idx,
                        is_first_card=(cards_processed + idx == 0),
                        page_error=f"tab_error: {type(e).__name__}",
                    )
                finally:
                    if page is not None:
                        try:
                            await browser_service.close_page(page)
                        except Exception:
                            pass

            tasks = [
                _process_tab(idx, card_id)
                for idx, card_id in enumerate(chunk)
            ]

            chunk_results = await asyncio.gather(*tasks, return_exceptions=True)

            for i, res in enumerate(chunk_results):
                if isinstance(res, CardDiagResult):
                    results.append(res)
                elif isinstance(res, BaseException):
                    results.append(CardDiagResult(
                        object_id=chunk[i] if i < len(chunk) else "?",
                        worker_idx=worker_idx,
                        page_error=f"gather_error: {type(res).__name__}",
                    ))

            cards_processed += len(chunk)
            await browser_service.close_all_pages()

            ok_count = sum(1 for r in results if r.api_success)
            err_count = sum(1 for r in results if r.page_error)
            logger.info(
                "воркер_прогресс",
                step=f"в={worker_idx}, {cards_processed}/{len(card_ids)}, "
                     f"ok={ok_count}, err={err_count}",
            )

            if chunk_start + max_tabs < len(card_ids):
                await asyncio.sleep(_INTER_CHUNK_PAUSE_SECONDS)

    except Exception as e:
        logger.error(
            "воркер_критическая_ошибка",
            error=str(e)[:200],
            error_type=type(e).__name__,
            step=f"в={worker_idx}",
        )
    finally:
        if launch_slot_held:
            _release_launch_slot()

        # НЕ останавливаем браузер здесь — возвращаем его для
        # централизованной остановки в main(). Это предотвращает
        # гонку exception_handler'ов при параллельном закрытии.

    ok_total = sum(1 for r in results if r.api_success)
    logger.info(
        "воркер_завершён",
        step=f"в={worker_idx}, всего={len(results)}, "
             f"ok={ok_total}, err={len(results) - ok_total}",
    )
    return (results, browser_service, warmup_info)


# ── Централизованная остановка браузеров ──


async def _stop_all_browsers(
    browsers: list[tuple[BrowserService, int]],
) -> None:
    """Останавливает все браузеры последовательными пачками.

    Playwright устанавливает exception_handler на event loop при
    каждом start() и восстанавливает при stop(). Параллельный вызов
    80 stop() создаёт гонку на loop.set_exception_handler() — один
    handler восстанавливает original, который на самом деле является
    handler'ом другого инстанса. Это может привести к зависанию
    playwright.stop() или к необработанным Future.

    Ограничение через семафор (_MAX_CONCURRENT_STOPS) гарантирует,
    что одновременно закрывается не более N Playwright-процессов.
    Последовательная остановка безопаснее, но слишком медленна
    при 80 браузерах (80 × 3с = 4 минуты). 5 одновременных — компромисс.

    Args:
        browsers: Список кортежей (BrowserService, worker_idx).
    """
    if not browsers:
        return

    total = len(browsers)
    logger.info(
        "централизованная_остановка_браузеров",
        step=f"всего={total}, параллельно={_MAX_CONCURRENT_STOPS}",
    )

    stop_semaphore = asyncio.Semaphore(_MAX_CONCURRENT_STOPS)
    stopped = 0

    async def _stop_one(browser_svc: BrowserService, w_idx: int) -> None:
        nonlocal stopped
        async with stop_semaphore:
            try:
                await asyncio.wait_for(
                    browser_svc.stop(),
                    timeout=15.0,
                )
            except asyncio.TimeoutError:
                logger.warning(
                    "остановка_браузера_таймаут",
                    step=f"в={w_idx}",
                )
            except Exception as e:
                logger.debug(
                    "ошибка_остановки_браузера",
                    error=str(e)[:100],
                    step=f"в={w_idx}",
                )
            finally:
                stopped += 1
                if stopped % 10 == 0 or stopped == total:
                    logger.info(
                        "прогресс_остановки_браузеров",
                        step=f"{stopped}/{total}",
                    )

    stop_tasks = [
        asyncio.create_task(
            _stop_one(bsvc, w_idx),
            name=f"stop-{w_idx}",
        )
        for bsvc, w_idx in browsers
    ]

    # Ждём завершения всех с жёстким таймаутом
    done, pending = await asyncio.wait(
        stop_tasks,
        timeout=_SHUTDOWN_PENDING_TIMEOUT + total * 0.5,
    )

    if pending:
        logger.warning(
            "остановка_не_завершилась_вовремя",
            step=f"завершено={len(done)}, зависло={len(pending)}",
        )
        for task in pending:
            task.cancel()
        # Даём отменённым задачам финализироваться
        await asyncio.wait(pending, timeout=3.0)

    logger.info(
        "все_браузеры_остановлены",
        step=f"всего={total}",
    )


# ── Очистка event loop от осиротевших задач ──


async def _cleanup_pending_tasks() -> None:
    """Отменяет все pending-задачи в event loop, кроме текущей.

    После параллельной остановки 80 Playwright-инстансов в event loop
    могут остаться незавершённые задачи (CDP-обработчики, pipe-читатели,
    callback'и от Node.js). asyncio.run() не завершится, пока эти задачи
    не будут отменены — это и вызывает зависание скрипта.
    """
    current_task = asyncio.current_task()
    all_tasks = asyncio.all_tasks()

    pending = [
        t for t in all_tasks
        if t is not current_task and not t.done()
    ]

    if not pending:
        return

    logger.info(
        "очистка_pending_задач",
        step=f"осталось={len(pending)}",
    )

    for task in pending:
        task.cancel()

    # Ждём завершения отменённых задач
    results = await asyncio.gather(*pending, return_exceptions=True)

    cancelled_count = sum(
        1 for r in results
        if isinstance(r, asyncio.CancelledError)
    )
    error_count = sum(
        1 for r in results
        if isinstance(r, BaseException) and not isinstance(r, asyncio.CancelledError)
    )

    logger.info(
        "pending_задачи_очищены",
        step=f"отменено={cancelled_count}, ошибок={error_count}",
    )


# ── Вывод сводной таблицы ──


def print_summary(
    all_results: list[CardDiagResult],
    elapsed: float,
    warmup_infos: list["WorkerWarmupInfo"] | None = None,
) -> None:
    """Выводит сводную таблицу диагностики в консоль.

    Args:
        all_results: Результаты диагностики по карточкам.
        elapsed: Общее время выполнения (секунды).
        warmup_infos: Результаты прогрева браузеров по воркерам
            (EnrichStrategies._warmup_browser).
    """
    warmup_infos = warmup_infos or []

    if warmup_infos:
        warmup_ok = sum(1 for w in warmup_infos if w.success)
        warmup_total = len(warmup_infos)
        warmup_times = [w.time_s for w in warmup_infos]

        sep_w = "=" * 72
        print(f"\n{sep_w}")
        print("  ПРОГРЕВ БРАУЗЕРОВ (EnrichStrategies._warmup_browser)")
        print(sep_w)
        print(
            f"\n  Успешно прогрето: {warmup_ok}/{warmup_total} "
            f"({warmup_ok / warmup_total * 100:.1f}%)"
        )
        if warmup_times:
            mn, mx = min(warmup_times), max(warmup_times)
            avg = statistics.mean(warmup_times)
            med = statistics.median(warmup_times)
            print(
                f"  Время прогрева: мин={mn:.1f}с, медиана={med:.1f}с, "
                f"среднее={avg:.1f}с, макс={mx:.1f}с"
            )

        failed_warmups = [w for w in warmup_infos if not w.success]
        if failed_warmups:
            print(f"\n  Воркеры, не прошедшие прогрев:")
            for w in failed_warmups:
                print(
                    f"    в={w.worker_idx}, прокси={w.proxy_label}, "
                    f"ошибка={w.error or 'неизвестна'}"
                )

    total = len(all_results)
    if total == 0:
        print("\n  Нет результатов для отображения (после прогрева не осталось карточек).\n")
        return

    page_ok = sum(1 for r in all_results if r.page_loaded)
    token_ok = sum(1 for r in all_results if r.token_captured)
    api_ok = sum(1 for r in all_results if r.api_success)

    page_times = [r.page_load_time_s for r in all_results if r.page_loaded]
    token_times = [r.token_wait_time_s for r in all_results if r.token_captured]
    api_times = [r.api_response_time_s for r in all_results if r.api_success]

    # Отдельно первые карточки воркера (обрабатываются сразу после прогрева)
    first_cards = [r for r in all_results if r.is_first_card]
    first_page_times = [r.page_load_time_s for r in first_cards if r.page_loaded]
    other_page_times = [
        r.page_load_time_s for r in all_results
        if r.page_loaded and not r.is_first_card
    ]

    error_reasons: dict[str, int] = {}
    for r in all_results:
        if r.page_error:
            error_reasons[r.page_error] = error_reasons.get(r.page_error, 0) + 1
        elif not r.token_captured and r.page_loaded:
            error_reasons["token_not_captured"] = (
                error_reasons.get("token_not_captured", 0) + 1
            )
        elif r.api_error:
            reason = r.api_error.split(":")[0] if ":" in r.api_error else r.api_error
            error_reasons[reason] = error_reasons.get(reason, 0) + 1

    busy_stats: dict[str, int] = {}
    for r in all_results:
        if r.api_busy_status:
            busy_stats[r.api_busy_status] = busy_stats.get(r.api_busy_status, 0) + 1

    worker_data: dict[int, WorkerStats] = {}
    for r in all_results:
        if r.worker_idx not in worker_data:
            worker_data[r.worker_idx] = WorkerStats(
                worker_idx=r.worker_idx, proxy=""
            )
        ws = worker_data[r.worker_idx]
        ws.cards_total += 1
        if r.page_loaded:
            ws.cards_page_ok += 1
            ws.page_load_times.append(r.page_load_time_s)
        if r.token_captured:
            ws.cards_token_ok += 1
            ws.token_wait_times.append(r.token_wait_time_s)
        if r.api_success:
            ws.cards_api_ok += 1
            ws.api_response_times.append(r.api_response_time_s)
        if r.page_error:
            ws.errors_by_reason[r.page_error] = (
                ws.errors_by_reason.get(r.page_error, 0) + 1
            )

    sep = "=" * 72

    print(f"\n{sep}")
    print("  ДИАГНОСТИКА ЭТАПА 2 — СВОДКА (с прогревом)")
    print(sep)

    print(f"\n  Общее время:         {elapsed:.1f} сек ({elapsed / 60:.1f} мин)")
    print(f"  Всего карточек:      {total}")
    print(f"  Страница загружена:  {page_ok}/{total} "
          f"({page_ok / total * 100:.1f}%)")
    print(f"  Токен перехвачен:    {token_ok}/{total} "
          f"({token_ok / total * 100:.1f}%)")
    print(f"  API-запрос успешен:  {api_ok}/{total} "
          f"({api_ok / total * 100:.1f}%)")

    def _stats_line(times: list[float], label: str) -> str:
        if not times:
            return f"  {label}: нет данных"
        mn = min(times)
        mx = max(times)
        avg = statistics.mean(times)
        med = statistics.median(times)
        p95 = sorted(times)[int(len(times) * 0.95)] if len(times) >= 20 else mx
        return (
            f"  {label}: "
            f"мин={mn:.2f}с, медиана={med:.2f}с, среднее={avg:.2f}с, "
            f"p95={p95:.2f}с, макс={mx:.2f}с"
        )

    print(f"\n{'─' * 72}")
    print("  ВРЕМЯ ОПЕРАЦИЙ")
    print(f"{'─' * 72}")
    print(_stats_line(page_times, "Загрузка страницы (все)    "))
    print(_stats_line(api_times, "API-запрос                 "))
    print(_stats_line(token_times, "Ожидание токена            "))

    # ── Сравнение первой карточки (сразу после прогрева) и остальных ──
    print(f"\n{'─' * 72}")
    print("  ПЕРВАЯ КАРТОЧКА (ПОСЛЕ ПРОГРЕВА) vs ОСТАЛЬНЫЕ")
    print(f"{'─' * 72}")
    print(_stats_line(first_page_times, "Первая карточка (после прогрева)"))
    print(_stats_line(other_page_times, "Остальные (тот же туннель)      "))

    first_ok = sum(1 for r in first_cards if r.page_loaded)
    first_total = len(first_cards)
    first_err = first_total - first_ok
    print(f"\n  Первых карточек: {first_total}, "
          f"загрузилось: {first_ok} ({first_ok / first_total * 100:.1f}%), "
          f"ошибок: {first_err}")

    if busy_stats:
        print(f"\n{'─' * 72}")
        print("  СТАТУС ЗАНЯТОСТИ (из успешных API-ответов)")
        print(f"{'─' * 72}")
        for status, count in sorted(busy_stats.items(), key=lambda x: -x[1]):
            print(f"  {status}: {count}")

    if error_reasons:
        print(f"\n{'─' * 72}")
        print("  ПРИЧИНЫ ОШИБОК")
        print(f"{'─' * 72}")
        for reason, count in sorted(error_reasons.items(), key=lambda x: -x[1]):
            pct = count / total * 100
            bar_len = min(int(pct / 2), 50)
            bar = "█" * bar_len + "░" * (50 - bar_len)
            print(f"  {count:>4} ({pct:>5.1f}%) {bar} {reason}")

    if len(worker_data) > 1:
        print(f"\n{'─' * 72}")
        print("  СТАТИСТИКА ПО ВОРКЕРАМ")
        print(f"{'─' * 72}")
        header = (
            f"  {'В':>3} {'Карт':>5} {'СтрОК':>6} {'ТокОК':>6} "
            f"{'ApiОК':>6} {'СтрМед':>7} {'ApiМед':>7} {'Ошиб':>5}"
        )
        print(header)
        print(f"  {'─' * 3} {'─' * 5} {'─' * 6} {'─' * 6} "
              f"{'─' * 6} {'─' * 7} {'─' * 7} {'─' * 5}")

        for ws in sorted(worker_data.values(), key=lambda x: x.worker_idx):
            page_med = (
                f"{statistics.median(ws.page_load_times):.1f}с"
                if ws.page_load_times else "—"
            )
            api_med = (
                f"{statistics.median(ws.api_response_times):.1f}с"
                if ws.api_response_times else "—"
            )
            err_count = ws.cards_total - ws.cards_api_ok
            print(
                f"  {ws.worker_idx:>3} {ws.cards_total:>5} "
                f"{ws.cards_page_ok:>6} {ws.cards_token_ok:>6} "
                f"{ws.cards_api_ok:>6} {page_med:>7} {api_med:>7} "
                f"{err_count:>5}"
            )

        worst_workers = sorted(
            worker_data.values(),
            key=lambda x: x.cards_total - x.cards_api_ok,
            reverse=True,
        )[:5]

        if worst_workers and any(w.errors_by_reason for w in worst_workers):
            print(f"\n  Топ-5 воркеров по ошибкам:")
            for ws in worst_workers:
                if ws.errors_by_reason:
                    top_err = max(
                        ws.errors_by_reason.items(), key=lambda x: x[1]
                    )
                    print(
                        f"    в={ws.worker_idx}: "
                        f"{ws.cards_total - ws.cards_api_ok} ошибок, "
                        f"главная: {top_err[0]} ({top_err[1]})"
                    )

    slowest = sorted(
        [r for r in all_results if r.page_loaded],
        key=lambda x: x.page_load_time_s,
        reverse=True,
    )[:10]

    if slowest:
        print(f"\n{'─' * 72}")
        print("  ТОП-10 САМЫХ МЕДЛЕННЫХ ЗАГРУЗОК СТРАНИЦ")
        print(f"{'─' * 72}")
        for r in slowest:
            token_s = (
                f"токен={r.token_wait_time_s:.1f}с"
                if r.token_captured else "токен=нет"
            )
            api_s = (
                f"api={r.api_response_time_s:.1f}с"
                if r.api_success
                else f"api_err={r.api_error[:25]}"
            )
            first_tag = " [1я]" if r.is_first_card else ""
            print(
                f"  id={r.object_id:>10} | стр={r.page_load_time_s:.1f}с"
                f" | {token_s} | {api_s} | в={r.worker_idx}{first_tag}"
            )

    print(f"\n{sep}\n")


# ── Точка входа ──


async def main() -> None:
    """Основная функция диагностики."""
    global _launch_semaphore, _launch_lock, _last_launch_time  # noqa: PLW0603

    configure(
        log_level="DEBUG",
        log_file_path="logs/test_enrichment_speed.log",
    )

    settings = Settings.load()

    # ── Загрузка ID объявлений из Excel-отчёта проекта ──
    ids_source_path = Path(
        os.getenv(_IDS_SOURCE_PATH_ENV, _DEFAULT_IDS_SOURCE_PATH)
    )
    if not ids_source_path.is_absolute():
        ids_source_path = PROJECT_ROOT / ids_source_path

    ids_source_column = os.getenv(
        _IDS_SOURCE_COLUMN_ENV, _DEFAULT_IDS_SOURCE_COLUMN
    )

    print(f"\n  Источник ID: {ids_source_path}")
    print(f"  Столбец с ID: '{ids_source_column}'")

    all_ids = load_test_ids_from_xlsx(ids_source_path, ids_source_column)
    print(f"  Загружено ID из файла: {len(all_ids)}")

    card_ids = all_ids[:MAX_TEST_CARDS] if MAX_TEST_CARDS > 0 else all_ids
    total_cards = len(card_ids)

    print(f"\n  Диагностика Этапа 2 (с прогревом, как в проекте): {total_cards} карточек")
    print(f"  Таймаут навигации: {settings.navigation_timeout} мс")
    print(f"  Ожидание токена: {_TOKEN_WAIT_SECONDS} сек")
    print(
        "  Прогрев: EnrichStrategies._warmup_browser "
        "(навигация на https://sutochno.ru + пауза, до 2 попыток)"
    )

    # ── Загрузка и проверка прокси ──
    proxies: list[ProxyConfig] = []
    proxy_service: ProxyService | None = None
    if settings.use_proxy:
        proxy_service = ProxyService(settings=settings)
        raw_proxies = proxy_service.load_proxies()
        print(f"  Загружено прокси: {len(raw_proxies)}")
        print(f"  Проверка прокси...")
        proxies = await proxy_service.check_proxies(raw_proxies)
        print(f"  Рабочих прокси: {len(proxies)}")

        if not proxies:
            print("  Нет рабочих прокси. Запуск без прокси.")

    # ── Конфигурация воркеров ──
    max_tabs = (
        _DIAG_MAX_TABS_OVERRIDE if _DIAG_MAX_TABS_OVERRIDE > 0 else settings.max_tabs
    )
    if _DIAG_MAX_TABS_OVERRIDE > 0 and _DIAG_MAX_TABS_OVERRIDE != settings.max_tabs:
        print(
            f"  Вкладок: {max_tabs} (оверрайд для теста; "
            f"в .env MAX_TABS={settings.max_tabs})"
        )

    if proxies:
        max_workers = min(len(proxies), settings.max_proxy_workers)
        active_proxies: list[ProxyConfig | None] = list(proxies[:max_workers])

        chunks: list[list[str]] = [[] for _ in range(max_workers)]
        for idx, card_id in enumerate(card_ids):
            chunks[idx % max_workers].append(card_id)

        print(f"  Воркеров: {max_workers}")
        print(f"  Вкладок на воркер: {max_tabs}")
        print(f"  Карточек на воркер: ~{total_cards // max_workers}")
    else:
        active_proxies = [None]
        chunks = [card_ids]
        max_workers = 1
        print(f"  Режим: без прокси, 1 воркер, {max_tabs} вкладок")

    # Полный список реальных ProxyConfig (без None) — передаётся в
    # _warmup_browser для исключения занятых прокси при поиске замены.
    all_proxy_configs: list[ProxyConfig] = [p for p in active_proxies if p is not None]

    # ── Резервируем стартовые прокси ДО запуска конкурентных воркеров ──
    # Без этого шага параллельный get_replacement_proxy из разных воркеров
    # мог бы выбрать прокси, уже статически отданную другому воркеру
    # (см. фикс гонки в proxy_service.py / enrich_strategies.py).
    if proxy_service is not None:
        for p in all_proxy_configs:
            await proxy_service.claim_proxy(p)

    print(f"  Одновременный запуск браузеров: {_MAX_CONCURRENT_BROWSER_LAUNCHES}")
    print(f"  Stagger между запусками: {_BROWSER_LAUNCH_STAGGER_SECONDS} сек")
    print()

    # ── Инициализация синхронизации ──
    _launch_semaphore = asyncio.Semaphore(_MAX_CONCURRENT_BROWSER_LAUNCHES)
    _launch_lock = asyncio.Lock()
    _last_launch_time = 0.0

    # ── Запуск всех воркеров сразу ──
    overall_start = time.perf_counter()

    worker_tasks: list[asyncio.Task] = []
    for i in range(max_workers):
        proxy = active_proxies[i] if i < len(active_proxies) else None
        chunk = chunks[i] if i < len(chunks) else []

        if not chunk:
            continue

        task = asyncio.create_task(
            run_worker(
                worker_idx=i + 1,
                card_ids=chunk,
                settings=settings,
                proxy=proxy,
                max_tabs=max_tabs,
                proxy_service=proxy_service,
                all_proxies=all_proxy_configs,
            ),
            name=f"diag-worker-{i + 1}",
        )
        worker_tasks.append(task)

    logger.info(
        "все_задачи_созданы",
        step=f"воркеров={len(worker_tasks)}, "
             f"семафор={_MAX_CONCURRENT_BROWSER_LAUNCHES}, "
             f"прогрев=включён (EnrichStrategies._warmup_browser)",
    )

    # ── Ожидание завершения ──
    all_results: list[CardDiagResult] = []
    browsers_to_stop: list[tuple[BrowserService, int]] = []
    warmup_infos: list[WorkerWarmupInfo] = []

    if worker_tasks:
        task_results = await asyncio.gather(*worker_tasks, return_exceptions=True)

        for i, res in enumerate(task_results):
            worker_idx = i + 1
            if isinstance(res, tuple) and len(res) == 3:
                card_results, browser_svc, warmup_info = res
                all_results.extend(card_results)
                browsers_to_stop.append((browser_svc, worker_idx))
                warmup_infos.append(warmup_info)
            elif isinstance(res, BaseException):
                logger.error(
                    "воркер_исключение",
                    error=str(res)[:200],
                    error_type=type(res).__name__,
                    step=f"в={worker_idx}",
                )

    overall_elapsed = time.perf_counter() - overall_start

    # ── Вывод сводки ПЕРЕД остановкой браузеров ──
    # Сводка выводится сразу — пользователь видит результаты,
    # пока браузеры останавливаются в фоне.
    print_summary(all_results, overall_elapsed, warmup_infos)

    # ── CSV ──
    csv_path = Path("data/diag_enrichment_speed_with_warmup.csv")
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    with open(csv_path, "w", encoding="utf-8") as f:
        f.write(
            "object_id,worker,is_first,page_loaded,page_time_s,page_error,"
            "token_captured,token_time_s,"
            "api_success,api_time_s,api_error,api_busy,total_time_s\n"
        )
        for r in all_results:
            pe = r.page_error.replace(",", ";").replace('"', "'")
            ae = r.api_error.replace(",", ";").replace('"', "'")
            f.write(
                f"{r.object_id},{r.worker_idx},{r.is_first_card},"
                f"{r.page_loaded},{r.page_load_time_s:.3f},{pe},"
                f"{r.token_captured},{r.token_wait_time_s:.3f},"
                f"{r.api_success},{r.api_response_time_s:.3f},{ae},"
                f"{r.api_busy_status},{r.total_time_s:.3f}\n"
            )

    warmup_csv_path = Path("data/diag_warmup_summary.csv")
    with open(warmup_csv_path, "w", encoding="utf-8") as f:
        f.write("worker,proxy,success,time_s,error\n")
        for w in warmup_infos:
            we = w.error.replace(",", ";").replace('"', "'")
            f.write(
                f"{w.worker_idx},{w.proxy_label},{w.success},"
                f"{w.time_s:.3f},{we}\n"
            )

    print(f"  Детальные результаты: {csv_path}")
    print(f"  Результаты прогрева: {warmup_csv_path}")
    print(f"  Логи: logs/test_enrichment_speed.log\n")

    # ── Централизованная остановка всех браузеров ──
    # Выполняется ПОСЛЕ вывода сводки и CSV — пользователь уже
    # получил результаты. Остановка может занять 30–60 секунд
    # при 80 браузерах, но это нормально.
    if browsers_to_stop:
        print(f"  Остановка {len(browsers_to_stop)} браузеров...")
        await _stop_all_browsers(browsers_to_stop)
        print("  Все браузеры остановлены.\n")

    # ── Очистка осиротевших задач в event loop ──
    # После закрытия 80 Playwright-инстансов в event loop могут
    # остаться pending-задачи (CDP-обработчики, pipe-читатели).
    # Без этой очистки asyncio.run() зависнет навсегда.
    await _cleanup_pending_tasks()


if __name__ == "__main__":
    asyncio.run(main())
