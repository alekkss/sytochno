"""Тестовый скрипт — диагностика запуска прокси-браузеров.

Воспроизводит логику Этапа 2 основной программы:
1. Загружает настройки из .env.
2. Загружает и проверяет прокси (как в __main__.py).
3. Читает первые 1000 ссылок из столбца "O" файла data/sutochno_report.xlsx.
4. Рассчитывает безопасное количество воркеров (MemoryMonitor).
5. Распределяет ссылки между рабочими прокси (distribute_listings).
6. Запускает воркеры — каждый открывает Chromium через прокси,
   прогревает браузер, затем обрабатывает свою порцию ссылок
   через MAX_TABS параллельных вкладок.
7. Каждая вкладка выполняет только навигацию (без обогащения).
8. Логирует каждый этап для диагностики.

Запуск:
    python -m scripts.test_proxy_browsers
"""

import asyncio
import sys
import time
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from src.config.logger import configure as configure_logging
from src.config.logger import get_logger
from src.config.settings import Settings
from src.models.proxy import ProxyConfig
from src.services.browser_service import BrowserService
from src.services.memory_monitor import MemoryMonitor
from src.services.proxy_service import ProxyService


# Максимум ссылок для теста
_MAX_LINKS: int = 1000

# Столбец "O" = 15-й столбец (содержит ссылки на объявления)
_LINK_COLUMN: int = 15

# Таймаут навигации для тестовой загрузки страницы (мс)
_TEST_NAVIGATION_TIMEOUT_MS: int = 30000

# Пауза после прогрева браузера (секунды)
_WARMUP_DELAY_SECONDS: float = 10.0

# Пауза между запуском вкладок внутри порции (секунды)
_TAB_START_DELAY_SECONDS: float = 1.0


@dataclass
class WorkerResult:
    """Результат работы одного воркера."""

    worker_idx: int
    proxy: str
    total_links: int
    successful_navigations: int
    failed_navigations: int
    duration_seconds: float
    error: str | None = None


def load_links_from_excel(excel_path: str, max_links: int) -> list[str]:
    """Загружает ссылки из столбца "O" Excel-файла.

    Читает файл в обычном режиме (не read_only), чтобы иметь доступ
    к гиперссылкам ячеек. Для 1000 строк это достаточно быстро.

    Args:
        excel_path: Путь к файлу sutochno_report.xlsx.
        max_links: Максимальное количество ссылок.

    Returns:
        Список URL-ссылок (без пустых значений).

    Raises:
        RuntimeError: Если файл не найден или столбец пуст.
    """
    path = Path(excel_path)
    if not path.exists():
        raise RuntimeError(
            f"Файл отчёта не найден: {excel_path}. "
            f"Сначала запустите основную программу для создания отчёта."
        )

    wb = load_workbook(filename=str(path), data_only=True)
    ws = wb.active

    if ws is None:
        wb.close()
        raise RuntimeError(f"Файл {excel_path} не содержит активного листа.")

    links: list[str] = []

    # Пропускаем заголовок (первая строка), читаем со второй
    for row_idx in range(2, ws.max_row + 1):
        if len(links) >= max_links:
            break

        cell = ws.cell(row=row_idx, column=_LINK_COLUMN)
        value = cell.value

        if value is None:
            continue

        # Приоритет: гиперссылка > текстовое значение ячейки
        link: str = ""

        if cell.hyperlink and cell.hyperlink.target:
            link = cell.hyperlink.target.strip()
        else:
            link = str(value).strip()

        if link and link.startswith("http"):
            links.append(link)

    wb.close()

    if not links:
        raise RuntimeError(
            f"Столбец 'O' в файле {excel_path} не содержит ссылок. "
            f"Убедитесь, что основная программа создала отчёт с данными."
        )

    return links


async def test_worker(
    settings: Settings,
    links: list[str],
    proxy: ProxyConfig,
    worker_idx: int,
    max_tabs: int,
) -> WorkerResult:
    """Воркер — запускает браузер через прокси и обрабатывает ссылки.

    Воспроизводит логику _worker из enrich_strategies.py:
    1. Создаёт BrowserService с прокси.
    2. Прогревает браузер (навигация на sutochno.ru + прокрутка + пауза).
    3. Обрабатывает ссылки порциями по max_tabs вкладок.
    4. Каждая вкладка выполняет только навигацию на страницу карточки.

    Args:
        settings: Настройки приложения.
        links: Порция ссылок для этого воркера.
        proxy: Прокси для этого воркера.
        worker_idx: Номер воркера (для логов).
        max_tabs: Количество параллельных вкладок.

    Returns:
        WorkerResult с результатами работы.
    """
    logger = get_logger("test_worker")

    if not links:
        return WorkerResult(
            worker_idx=worker_idx,
            proxy=str(proxy),
            total_links=0,
            successful_navigations=0,
            failed_navigations=0,
            duration_seconds=0.0,
        )

    browser_service = BrowserService(settings=settings)
    worker_start = time.perf_counter()
    successful = 0
    failed = 0

    try:
        # Шаг 1: Запуск браузера с прокси
        logger.info(
            "воркер_запуск_браузера",
            step=f"воркер={worker_idx}, прокси={proxy}",
        )
        await browser_service.start(proxy=proxy)

        logger.info(
            "воркер_браузер_запущен",
            step=f"воркер={worker_idx}",
        )

        # Шаг 2: Прогрев — навигация на главную + прокрутка + пауза
        logger.info(
            "воркер_прогрев",
            step=f"воркер={worker_idx}",
        )
        await browser_service.navigate("https://sutochno.ru")
        await browser_service.scroll_page()
        await asyncio.sleep(_WARMUP_DELAY_SECONDS)

        logger.info(
            "воркер_прогрет",
            step=f"воркер={worker_idx}, ссылок={len(links)}",
        )

        # Шаг 3: Обработка ссылок порциями по max_tabs
        for chunk_start in range(0, len(links), max_tabs):
            chunk = links[chunk_start: chunk_start + max_tabs]

            # Создаём задачи для параллельных вкладок
            tasks = [
                _navigate_in_tab(
                    browser_service, link, worker_idx, tab_idx
                )
                for tab_idx, link in enumerate(chunk)
            ]

            results = await asyncio.gather(*tasks, return_exceptions=True)

            # Подсчитываем результаты
            for result in results:
                if isinstance(result, Exception):
                    failed += 1
                elif result is True:
                    successful += 1
                else:
                    failed += 1

            # Закрываем дополнительные вкладки после каждой порции
            await browser_service.close_all_pages()

            processed = chunk_start + len(chunk)
            logger.info(
                "воркер_прогресс",
                step=f"воркер={worker_idx}, "
                     f"обработано={processed}/{len(links)}, "
                     f"успех={successful}, ошибок={failed}",
            )

    except Exception as e:
        logger.error(
            "воркер_критическая_ошибка",
            error=str(e),
            error_type=type(e).__name__,
            step=f"воркер={worker_idx}",
        )
        elapsed = time.perf_counter() - worker_start
        return WorkerResult(
            worker_idx=worker_idx,
            proxy=str(proxy),
            total_links=len(links),
            successful_navigations=successful,
            failed_navigations=failed,
            duration_seconds=elapsed,
            error=str(e),
        )
    finally:
        # Гарантированная остановка браузера
        try:
            await browser_service.stop()
        except Exception as e:
            logger.warning(
                "воркер_ошибка_остановки",
                error=str(e),
                step=f"воркер={worker_idx}",
            )

    elapsed = time.perf_counter() - worker_start

    logger.info(
        "воркер_завершён",
        step=f"воркер={worker_idx}, "
             f"успех={successful}, ошибок={failed}, "
             f"время={elapsed:.1f}с",
    )

    return WorkerResult(
        worker_idx=worker_idx,
        proxy=str(proxy),
        total_links=len(links),
        successful_navigations=successful,
        failed_navigations=failed,
        duration_seconds=elapsed,
    )


async def _navigate_in_tab(
    browser_service: BrowserService,
    link: str,
    worker_idx: int,
    tab_idx: int,
) -> bool:
    """Открывает вкладку и выполняет навигацию на указанную ссылку.

    Воспроизводит логику _process_one_tab из enrich_strategies.py:
    - Задержка перед стартом (кроме первой вкладки).
    - Создание вкладки.
    - Навигация.
    - Закрытие вкладки.

    Args:
        browser_service: Сервис браузера.
        link: URL для навигации.
        worker_idx: Номер воркера (для логов).
        tab_idx: Номер вкладки в порции (для задержки).

    Returns:
        True если навигация успешна, False при ошибке.
    """
    logger = get_logger("test_tab")
    page = None

    # Задержка между запуском вкладок (как в основной программе)
    if tab_idx > 0:
        await asyncio.sleep(_TAB_START_DELAY_SECONDS)

    try:
        page = await browser_service.create_page()
        page.set_default_navigation_timeout(_TEST_NAVIGATION_TIMEOUT_MS)

        await page.goto(link, wait_until="domcontentloaded")

        # Проверяем, что страница загрузилась (есть контент)
        content = await page.content()
        if len(content) < 500:
            logger.debug(
                "вкладка_пустой_контент",
                step=f"воркер={worker_idx}, вкладка={tab_idx}, url={link[:60]}",
            )
            return False

        return True

    except Exception as e:
        logger.debug(
            "вкладка_ошибка_навигации",
            error=str(e)[:100],
            error_type=type(e).__name__,
            step=f"воркер={worker_idx}, вкладка={tab_idx}",
        )
        return False

    finally:
        if page is not None:
            try:
                if not page.is_closed():
                    await page.close()
            except Exception:
                pass


def distribute_links(links: list[str], worker_count: int) -> list[list[str]]:
    """Распределяет ссылки поровну между воркерами.

    Точная копия логики ProxyService.distribute_listings.

    Args:
        links: Общий список ссылок.
        worker_count: Количество воркеров.

    Returns:
        Список списков — порция ссылок для каждого воркера.
    """
    if worker_count <= 0:
        return [links]

    chunks: list[list[str]] = [[] for _ in range(worker_count)]

    for idx, link in enumerate(links):
        chunk_idx = idx % worker_count
        chunks[chunk_idx].append(link)

    return chunks


async def run() -> None:
    """Основная логика тестового скрипта."""
    # --- Шаг 1: Загрузка настроек ---
    try:
        settings = Settings.load()
    except RuntimeError as e:
        print(f"[ОШИБКА] {e}", file=sys.stderr)  # noqa: T201
        sys.exit(1)

    # --- Шаг 2: Конфигурация логирования ---
    configure_logging(
        log_level=settings.log_level,
        log_file_path="logs/test_proxy_browsers.log",
    )
    logger = get_logger("test_proxy_browsers")

    logger.info("=" * 70)
    logger.info(
        "ТЕСТ_ЗАПУСКА_ПРОКСИ_БРАУЗЕРОВ",
        step="начало",
    )
    logger.info("=" * 70)

    # --- Шаг 3: Параметры из .env ---
    logger.info(
        "параметры_из_env",
        step=f"USE_PROXY={settings.use_proxy}, "
             f"MAX_PROXY_WORKERS={settings.max_proxy_workers}, "
             f"MAX_TABS={settings.max_tabs}, "
             f"TAB_DELAY_MS={settings.tab_delay_ms}, "
             f"MEMORY_LIMIT_MB={settings.memory_limit_mb}, "
             f"HEADLESS_MODE={settings.headless_mode}",
    )

    # --- Шаг 4: Проверка USE_PROXY ---
    if not settings.use_proxy:
        logger.error(
            "USE_PROXY_выключен",
            step="USE_PROXY=false в .env — тест прокси-браузеров невозможен. "
                 "Установите USE_PROXY=true и повторите.",
        )
        sys.exit(1)

    # --- Шаг 5: Загрузка ссылок из Excel ---
    excel_path = settings.export_path
    logger.info(
        "загрузка_ссылок_из_excel",
        step=f"файл={excel_path}, макс_ссылок={_MAX_LINKS}",
    )

    try:
        links = load_links_from_excel(excel_path, _MAX_LINKS)
    except RuntimeError as e:
        logger.error("ошибка_загрузки_ссылок", error=str(e))
        sys.exit(1)

    logger.info(
        "ссылки_загружены",
        total=len(links),
        step=f"первая={links[0][:60]}...",
    )

    # --- Шаг 6: Загрузка и проверка прокси ---
    logger.info("загрузка_прокси", step="начало")
    proxy_service = ProxyService(settings=settings)

    try:
        all_proxies = proxy_service.load_proxies()
    except RuntimeError as e:
        logger.error("ошибка_загрузки_прокси", error=str(e))
        sys.exit(1)

    logger.info(
        "прокси_загружены_из_файла",
        total=len(all_proxies),
        step=f"файл={settings.proxies_path}",
    )

    logger.info("начало_проверки_прокси", total=len(all_proxies))
    working_proxies = await proxy_service.check_proxies(all_proxies)

    logger.info(
        "результат_проверки_прокси",
        step=f"всего={len(all_proxies)}, рабочих={len(working_proxies)}, "
             f"нерабочих={len(all_proxies) - len(working_proxies)}",
    )

    if not working_proxies:
        logger.error(
            "нет_рабочих_прокси",
            step="Ни одна прокси не прошла проверку. Тест невозможен.",
        )
        sys.exit(1)

    # --- Шаг 7: Расчёт количества воркеров (как в enrich_listings_parallel) ---
    max_workers = settings.max_proxy_workers
    proxies_to_use = working_proxies[:max_workers]

    logger.info(
        "ограничение_по_max_proxy_workers",
        step=f"рабочих_прокси={len(working_proxies)}, "
             f"MAX_PROXY_WORKERS={max_workers}, "
             f"к_использованию={len(proxies_to_use)}",
    )

    # Расчёт через MemoryMonitor (точная копия из enrich_listings_parallel)
    memory_monitor = MemoryMonitor(
        memory_limit_mb=settings.memory_limit_mb,
        max_tabs=settings.max_tabs,
    )

    requested_workers = len(proxies_to_use)
    safe_workers = memory_monitor.calculate_safe_workers(requested_workers)

    logger.info(
        "расчёт_безопасных_воркеров",
        step=f"запрошено={requested_workers}, "
             f"безопасно={safe_workers}, "
             f"memory_limit_mb={settings.memory_limit_mb}, "
             f"на_воркер_мб={memory_monitor.estimate_worker_mb()}",
    )

    # Итоговое количество воркеров
    active_proxies = proxies_to_use[:safe_workers]

    logger.info(
        "ИТОГО_ВОРКЕРОВ_К_ЗАПУСКУ",
        total=len(active_proxies),
        step=f"из {len(all_proxies)} прокси в файле → "
             f"{len(working_proxies)} рабочих → "
             f"{len(proxies_to_use)} после MAX_PROXY_WORKERS → "
             f"{len(active_proxies)} после MemoryMonitor",
    )

    # --- Шаг 8: Распределение ссылок между воркерами ---
    chunks = distribute_links(links, len(active_proxies))

    logger.info("распределение_ссылок")
    for idx, chunk in enumerate(chunks, start=1):
        logger.info(
            "порция_воркера",
            step=f"воркер={idx}, ссылок={len(chunk)}, прокси={active_proxies[idx-1]}",
        )

    # --- Шаг 9: Запуск всех воркеров параллельно ---
    logger.info("=" * 70)
    logger.info(
        "ЗАПУСК_ВОРКЕРОВ",
        total=len(active_proxies),
        step=f"ссылок={len(links)}, вкладок_на_воркер={settings.max_tabs}",
    )
    logger.info("=" * 70)

    parallel_start = time.perf_counter()

    tasks = [
        test_worker(
            settings=settings,
            links=chunk,
            proxy=proxy,
            worker_idx=idx,
            max_tabs=settings.max_tabs,
        )
        for idx, (chunk, proxy) in enumerate(
            zip(chunks, active_proxies), start=1
        )
    ]

    results = await asyncio.gather(*tasks, return_exceptions=True)

    parallel_elapsed = time.perf_counter() - parallel_start

    # --- Шаг 10: Сводка результатов ---
    logger.info("=" * 70)
    logger.info("СВОДКА_РЕЗУЛЬТАТОВ", step="итого")
    logger.info("=" * 70)

    successful_workers = 0
    failed_workers = 0
    total_successful_navs = 0
    total_failed_navs = 0

    for result in results:
        if isinstance(result, Exception):
            failed_workers += 1
            logger.error(
                "воркер_исключение",
                error=str(result),
                error_type=type(result).__name__,
            )
        elif isinstance(result, WorkerResult):
            if result.error:
                failed_workers += 1
                logger.warning(
                    "воркер_с_ошибкой",
                    step=f"воркер={result.worker_idx}, "
                         f"прокси={result.proxy}, "
                         f"ошибка={result.error}",
                )
            else:
                successful_workers += 1

            total_successful_navs += result.successful_navigations
            total_failed_navs += result.failed_navigations

            logger.info(
                "результат_воркера",
                step=f"воркер={result.worker_idx}, "
                     f"прокси={result.proxy}, "
                     f"ссылок={result.total_links}, "
                     f"успех={result.successful_navigations}, "
                     f"ошибок={result.failed_navigations}, "
                     f"время={result.duration_seconds:.1f}с"
                     f"{', ОШИБКА: ' + result.error if result.error else ''}",
            )

    logger.info("─" * 70)
    logger.info(
        "ИТОГО",
        step=f"воркеров_запущено={len(active_proxies)}, "
             f"успешных={successful_workers}, "
             f"с_ошибками={failed_workers}",
        total=f"навигаций_успех={total_successful_navs}, "
              f"навигаций_ошибок={total_failed_navs}, "
              f"общее_время={parallel_elapsed:.1f}с",
    )
    logger.info("─" * 70)

    # Вывод в консоль для удобства
    print("\n" + "=" * 70)  # noqa: T201
    print("РЕЗУЛЬТАТЫ ТЕСТА ПРОКСИ-БРАУЗЕРОВ")  # noqa: T201
    print("=" * 70)  # noqa: T201
    print(f"  Прокси в файле:          {len(all_proxies)}")  # noqa: T201
    print(f"  Прокси рабочих:          {len(working_proxies)}")  # noqa: T201
    print(f"  MAX_PROXY_WORKERS:       {max_workers}")  # noqa: T201
    print(f"  После MemoryMonitor:     {safe_workers}")  # noqa: T201
    print(f"  Воркеров запущено:       {len(active_proxies)}")  # noqa: T201
    print(f"  Воркеров успешных:       {successful_workers}")  # noqa: T201
    print(f"  Воркеров с ошибками:     {failed_workers}")  # noqa: T201
    print(f"  MAX_TABS на воркер:      {settings.max_tabs}")  # noqa: T201
    print(f"  Ссылок обработано:       {total_successful_navs + total_failed_navs}")  # noqa: T201
    print(f"  Навигаций успешных:      {total_successful_navs}")  # noqa: T201
    print(f"  Навигаций с ошибками:    {total_failed_navs}")  # noqa: T201
    print(f"  Общее время:             {parallel_elapsed:.1f}с")  # noqa: T201
    print("=" * 70)  # noqa: T201

    if len(active_proxies) < len(all_proxies):
        print("\n⚠️  ВНИМАНИЕ: Количество воркеров меньше количества прокси в файле!")  # noqa: T201
        print("   Возможные причины ограничения:")  # noqa: T201

        if len(working_proxies) < len(all_proxies):
            print(  # noqa: T201
                f"   → {len(all_proxies) - len(working_proxies)} прокси не прошли проверку "
                f"(нерабочие/заблокированные)"
            )

        if len(working_proxies) > max_workers:
            print(  # noqa: T201
                f"   → MAX_PROXY_WORKERS={max_workers} ограничивает количество "
                f"(рабочих прокси {len(working_proxies)} > лимита {max_workers})"
            )

        if safe_workers < requested_workers:
            print(  # noqa: T201
                f"   → MemoryMonitor ограничил с {requested_workers} до {safe_workers} "
                f"(MEMORY_LIMIT_MB={settings.memory_limit_mb})"
            )


def main() -> None:
    """Синхронная точка входа."""
    asyncio.run(run())


if __name__ == "__main__":
    main()
