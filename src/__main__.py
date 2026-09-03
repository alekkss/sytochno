"""Точка входа приложения — сборка зависимостей и запуск pipeline.

Схема прогона (начиная с версии пула):
  Этап 1 (сбор каталога) выполняется по расписанию — слоты МСК из
  CATALOG_SYNC_TIMES (по умолчанию 01:00 и 19:00). Собранные ID
  синхронизируются с пулом (таблица listing_pool): новые добавляются,
  метки last_seen обновляются. Удалений нет.

  Каждый прогон начинает этап 2 (batch-обогащение) с ID из пула:
  - в прогоне с каталогом обрабатываются свежесобранные листинги;
  - в прогоне без каталога листинги загружаются из БД по пулу
    (ID пула без записи в listings пропускаются).

  Токен для однопоточного batch-обогащения без каталога получается
  через отдельный браузер (загрузка страницы поиска + перехват).
"""

import asyncio
import json
import signal
import sys
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path

from src.config.logger import configure as configure_logging
from src.config.logger import get_logger
from src.config.settings import Settings
from src.models.booking_event import AnyEvent
from src.models.listing import RawListing
from src.models.proxy import ProxyConfig
from src.repositories.base import BaseListingRepository
from src.repositories.base_comparison_events_repository import (
    BaseComparisonEventsRepository,
)
from src.repositories.db_factory import RepositoryPair, create_repositories
from src.repositories.postgres_comparison_events_repository import (
    PostgreSQLComparisonEventsRepository,
)
from src.repositories.snapshot_repository import BaseSnapshotRepository
from src.services.browser_service import BrowserService
from src.services.catalog_schedule import (
    CATALOG_SYNC_STATE_FILENAME,
    CatalogSchedule,
)
from src.services.comparison_export_service import ComparisonExportService
from src.services.comparison_service import ComparisonService
from src.services.data_cleaner_service import DataCleanerService
from src.services.export_service import ExportService
from src.services.listing.batch_enrichment_service import BatchEnrichmentService
from src.services.pool_service import PoolService
from src.services.proxy_service import ProxyService
from src.services.scraper_service import ScraperService
from src.services.snapshot_service import SnapshotService


# ── Константы цикла ──────────────────────────────────────────

# Пауза между прогонами (секунды)
_CYCLE_PAUSE_SECONDS: int = 120

# Путь к JSON-файлу статуса прогона (относительно data/)
_RUN_STATUS_FILENAME: str = "last_run_status.json"

# Часовой пояс Москвы (UTC+3) — для расчёта окна паузы
_MSK_OFFSET: timedelta = timedelta(hours=3)
_MSK_TZ: timezone = timezone(_MSK_OFFSET)

# Флаг остановки цикла (устанавливается через SIGTERM/SIGINT)
_shutdown_requested: bool = False


def _request_shutdown(signum: int, frame: "any") -> None:  # type: ignore[name-defined]
    """Обработчик сигналов SIGTERM и SIGINT — запрашивает остановку цикла.

    Устанавливает глобальный флаг, который проверяется в паузе между прогонами
    и перед стартом нового прогона. Текущий прогон не прерывается — он
    завершается штатно, после чего цикл останавливается.

    Args:
        signum: Номер сигнала.
        frame: Текущий стек-фрейм (не используется).
    """
    global _shutdown_requested  # noqa: PLW0603
    _shutdown_requested = True


def _is_in_pause_window(
    pause_start: tuple[int, int],
    pause_end: tuple[int, int],
) -> bool:
    """Проверяет, попадает ли текущее время МСК в окно паузы.

    Поддерживает окна, пересекающие полночь (например, 22:50–00:10).

    Args:
        pause_start: Начало паузы (часы, минуты) по МСК.
        pause_end: Конец паузы (часы, минуты) по МСК.

    Returns:
        True если текущее время находится внутри окна паузы.
    """
    now_msk = datetime.now(_MSK_TZ)
    current_minutes = now_msk.hour * 60 + now_msk.minute

    start_minutes = pause_start[0] * 60 + pause_start[1]
    end_minutes = pause_end[0] * 60 + pause_end[1]

    if start_minutes <= end_minutes:
        # Обычное окно (например, 02:00–06:00)
        return start_minutes <= current_minutes < end_minutes
    else:
        # Окно пересекает полночь (например, 22:50–00:10)
        return current_minutes >= start_minutes or current_minutes < end_minutes


def _seconds_until_pause_end(pause_end: tuple[int, int]) -> int:
    """Вычисляет количество секунд до конца паузы.

    Args:
        pause_end: Конец паузы (часы, минуты) по МСК.

    Returns:
        Количество секунд до момента pause_end. Если pause_end уже
        «в прошлом» сегодня (окно через полночь) — считает до завтра.
    """
    now_msk = datetime.now(_MSK_TZ)

    # Целевое время сегодня
    target_today = now_msk.replace(
        hour=pause_end[0],
        minute=pause_end[1],
        second=0,
        microsecond=0,
    )

    if target_today > now_msk:
        delta = target_today - now_msk
    else:
        # Конец паузы уже прошёл сегодня — значит это «завтра»
        target_tomorrow = target_today + timedelta(days=1)
        delta = target_tomorrow - now_msk

    return int(delta.total_seconds())


async def _wait_for_pause_end(
    pause_start: tuple[int, int],
    pause_end: tuple[int, int],
    logger: "any",  # type: ignore[name-defined]
) -> None:
    """Ожидает окончания паузы, если текущее время попадает в окно.

    Проверяет SIGTERM каждую секунду — при запросе остановки
    выходит немедленно, не дожидаясь конца паузы.

    Args:
        pause_start: Начало паузы (часы, минуты) по МСК.
        pause_end: Конец паузы (часы, минуты) по МСК.
        logger: Логгер для записи состояния.
    """
    global _shutdown_requested

    if not _is_in_pause_window(pause_start, pause_end):
        return

    wait_seconds = _seconds_until_pause_end(pause_end)
    end_time_str = f"{pause_end[0]:02d}:{pause_end[1]:02d}"

    logger.info(
        "ночная_пауза_начата",
        step=f"окно={pause_start[0]:02d}:{pause_start[1]:02d}–{end_time_str} МСК, "
             f"ожидание≈{wait_seconds // 60}мин",
    )

    # Ожидаем с проверкой SIGTERM каждую секунду
    elapsed = 0
    while elapsed < wait_seconds:
        if _shutdown_requested:
            logger.info(
                "остановка_запрошена_во_время_ночной_паузы",
                step=f"ожидание_прервано_на={elapsed}с",
            )
            return

        # Перепроверяем: вдруг время уже вышло из окна паузы
        if not _is_in_pause_window(pause_start, pause_end):
            break

        await asyncio.sleep(1.0)
        elapsed += 1

    logger.info(
        "ночная_пауза_завершена",
        step=f"ожидание={elapsed}с",
    )


def _write_run_status(
    data_dir: str,
    status: str,
    started_at: str,
    finished_at: str,
    listings_count: int = 0,
    events_count: int = 0,
    error: str | None = None,
    run_number: int = 0,
) -> None:
    """Записывает результат прогона в JSON-файл для чтения админкой.

    Файл перезаписывается при каждом прогоне. Админка мониторит
    изменение файла (по полю run_number или finished_at) и создаёт
    запись в истории запусков.

    Args:
        data_dir: Путь к папке data/ парсера.
        status: Статус прогона ("success", "failed", "cancelled").
        started_at: Время начала прогона (ISO 8601, UTC).
        finished_at: Время завершения прогона (ISO 8601, UTC).
        listings_count: Количество объявлений в БД после прогона.
        events_count: Количество событий (брони + отмены).
        error: Текст ошибки (только для status="failed").
        run_number: Порядковый номер прогона в текущей сессии.
    """
    status_path = Path(data_dir) / _RUN_STATUS_FILENAME

    data = {
        "status": status,
        "started_at": started_at,
        "finished_at": finished_at,
        "listings_count": listings_count,
        "events_count": events_count,
        "error": error,
        "run_number": run_number,
    }

    try:
        status_path.parent.mkdir(parents=True, exist_ok=True)
        status_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except OSError:
        # Не критично — админка просто не увидит этот прогон
        pass


def _count_enriched(listings: list) -> int:
    """Подсчитывает количество обогащённых карточек.

    Args:
        listings: Список карточек.

    Returns:
        Количество карточек с данными календаря или цен.
    """
    count = 0
    for listing in listings:
        if listing.enrichment_skip_reason is not None:
            continue
        if listing.calendar_60_days and any(c == 1 for c in listing.calendar_60_days):
            count += 1
        elif listing.prices_60_days and any(p > 0 for p in listing.prices_60_days):
            count += 1
    return count


def _count_unenriched(listings: list) -> int:
    """Подсчитывает количество необогащённых карточек (без фатальных).

    Args:
        listings: Список карточек.

    Returns:
        Количество карточек без данных и без фатальной причины.
    """
    count = 0
    for listing in listings:
        if listing.enrichment_skip_reason is not None:
            continue
        has_calendar = listing.calendar_60_days and any(
            c == 1 for c in listing.calendar_60_days
        )
        has_prices = listing.prices_60_days and any(
            p > 0 for p in listing.prices_60_days
        )
        if not has_calendar and not has_prices:
            count += 1
    return count


def _get_unenriched_listings(listings: list) -> list:
    """Возвращает список необогащённых карточек без фатальных причин.

    Args:
        listings: Полный список карточек.

    Returns:
        Карточки, у которых нет данных календаря/цен и нет enrichment_skip_reason.
    """
    return [
        l for l in listings
        if l.enrichment_skip_reason is None
        and not (l.calendar_60_days and any(c == 1 for c in l.calendar_60_days))
        and not (l.prices_60_days and any(p > 0 for p in l.prices_60_days))
    ]


def _build_events_repository(
    settings: Settings,
    logger: "any",  # type: ignore[name-defined]
) -> BaseComparisonEventsRepository | None:
    """Создаёт репозиторий событий сравнения для записи в БД.

    Репозиторий создаётся только при DB_TYPE=postgresql. При любой
    ошибке инициализации (нет драйвера, БД недоступна, таблица не
    создана) — возвращает None. Экспорт в Excel продолжит работать
    без записи в БД.

    Args:
        settings: Настройки приложения.
        logger: Логгер для сообщений о результате инициализации.

    Returns:
        Инициализированный репозиторий или None, если БД недоступна
        или используется SQLite.
    """
    if settings.db_type != "postgresql":
        logger.info(
            "запись_событий_в_бд_отключена_sqlite",
            step="events_repo",
        )
        return None

    dsn = (
        f"host={settings.pg_host} "
        f"port={settings.pg_port} "
        f"dbname={settings.pg_name} "
        f"user={settings.pg_user} "
        f"password={settings.pg_password}"
    )

    try:
        repo = PostgreSQLComparisonEventsRepository(dsn=dsn)
        repo.initialize()
        logger.info(
            "репозиторий_событий_готов",
            step="events_repo",
        )
        return repo
    except Exception as e:
        logger.warning(
            "ошибка_инициализации_репозитория_событий",
            error=str(e)[:300],
            error_type=type(e).__name__,
            step="events_repo",
        )
        return None


def _load_pool_listings(
    pool_ids: list[str],
    repository: BaseListingRepository,
) -> tuple[list[RawListing], int]:
    """Загружает из БД листинги, входящие в пул.

    Используется в прогонах без каталога: этап batch-обогащения
    работает с карточками из базы, отфильтрованными по ID пула.
    ID пула, для которых в listings нет записи, пропускаются —
    их метаданные появятся после ближайшей синхронизации каталога.

    Args:
        pool_ids: Все ID пула.
        repository: Репозиторий объявлений.

    Returns:
        Кортеж (листинги из пула, количество пропущенных ID
        без записи в БД).
    """
    pool_set = set(pool_ids)
    all_listings = repository.get_all()

    result = [l for l in all_listings if l.external_id in pool_set]
    missing_count = len(pool_set) - len(result)

    return result, missing_count


async def _run_standalone_batch(
    settings: Settings,
    batch_enrichment_service: BatchEnrichmentService,
    listings: list,
    logger: "any",  # type: ignore[name-defined]
    purpose: str = "batch",
) -> bool:
    """Выполняет batch-обогащение через отдельный браузер без прокси.

    Запускает один браузер, загружает страницу поиска, перехватывает
    токен API и передаёт карточки в enrich_batch (bulk + скользящее
    окно). Используется в двух сценариях:
    - основной batch в прогоне без каталога (нет токена этапа 1);
    - однократный batch-retry необработанных карточек.

    Args:
        settings: Настройки приложения.
        batch_enrichment_service: Сервис batch-обогащения.
        listings: Карточки для обогащения (обогащаются in-place).
        logger: Логгер.
        purpose: Метка сценария для логов ("batch" или "batch_retry").

    Returns:
        True если токен получен и batch выполнен, False при неудаче.
    """
    browser_service = BrowserService(settings=settings)

    try:
        await browser_service.start()

        page = browser_service.page
        search_url = settings.search_urls[0]

        logger.info(
            "загрузка_страницы_поиска_для_токена",
            purpose=purpose,
            step=f"url={search_url[:80]}...",
        )

        # ── Перехват токена со страницы поиска ──
        captured_token: list[str] = []

        async def _intercept(route, request):
            url = request.url
            if "sutochno.ru/api/json" in url and not captured_token:
                token = (
                    request.headers.get("token")
                    or request.headers.get("Token")
                )
                if token:
                    captured_token.append(token)
            try:
                await route.continue_()
            except Exception as e:
                if "Route is already handled" not in str(e):
                    raise

        await page.route("**/api/json/**", _intercept)

        await page.goto(
            search_url,
            wait_until="domcontentloaded",
            timeout=settings.navigation_timeout,
        )

        # Ожидание перехвата токена (до 20 секунд)
        elapsed = 0.0
        while elapsed < 20.0:
            if captured_token:
                break
            await asyncio.sleep(0.5)
            elapsed += 0.5

        # Снимаем route handler — критически важно для последующих fetch()
        try:
            await page.unroute("**/api/json/**")
        except Exception:
            pass

        # Стабилизация сессии
        await asyncio.sleep(3.0)

        if not captured_token:
            logger.warning(
                "токен_не_перехвачен",
                purpose=purpose,
                step=f"ожидание={elapsed:.1f}с",
            )
            return False

        token = captured_token[0]

        logger.info(
            "токен_получен",
            purpose=purpose,
            step=f"ожидание={elapsed:.1f}с, карточек={len(listings)}",
        )

        # ── Batch-обогащение карточек ──
        await batch_enrichment_service.enrich_batch(
            page=page,
            token=token,
            listings=listings,
            search_url=search_url,
        )

        return True

    except Exception as e:
        logger.warning(
            "ошибка_standalone_batch",
            purpose=purpose,
            error=str(e)[:300],
            error_type=type(e).__name__,
        )
        return False
    finally:
        try:
            await browser_service.stop()
        except Exception:
            pass


async def run() -> None:
    """Один прогон pipeline — каталог по расписанию, обогащение из пула.

    Возвращает управление после завершения. Не содержит цикла —
    цикл реализован в run_loop(). Это позволяет корректно обрабатывать
    критические ошибки на уровне цикла.

    Raises:
        RuntimeError: При ошибке загрузки конфигурации.
        Exception: Любая критическая ошибка pipeline.
    """
    # --- Шаг 1: Загрузка конфигурации ---
    settings = Settings.load()

    # --- Шаг 2: Конфигурация логирования ---
    configure_logging(
        log_level=settings.log_level,
        log_file_path=settings.log_file_path,
    )
    logger = get_logger("main")
    logger.info(
        "прогон_начат",
        step="init",
        search_urls_count=len(settings.search_urls),
        max_pages=settings.max_pages,
        db_type=settings.db_type,
    )

    data_dir = str(Path(settings.export_path).parent)

    # --- Шаг 3: Инициализация репозиториев через фабрику ---
    repos = create_repositories(settings)
    repository = repos.listing
    snapshot_repository = repos.snapshot
    pool_repository = repos.pool

    # --- Шаг 3.1: Инициализация репозитория событий сравнения (PostgreSQL) ---
    events_repository = _build_events_repository(settings, logger)

    # --- Шаг 4: Загрузка и проверка прокси ---
    working_proxies: list[ProxyConfig] = []
    proxy_service: ProxyService | None = None

    if settings.use_proxy:
        logger.info("загрузка_прокси", step="proxy")
        proxy_service = ProxyService(settings=settings)

        try:
            proxies = proxy_service.load_proxies()
            working_proxies = await proxy_service.check_proxies(proxies)
        except RuntimeError as e:
            logger.warning(
                "ошибка_загрузки_прокси",
                error=str(e),
                step="proxy",
            )

        if working_proxies:
            logger.info(
                "прокси_готовы",
                total=len(working_proxies),
                step="proxy",
            )
        else:
            logger.warning(
                "рабочих_прокси_нет_работа_без_прокси",
                step="proxy",
            )

    # --- Шаг 5: Создание сервисов ---
    browser_service = BrowserService(settings=settings)
    scraper_service = ScraperService(
        settings=settings,
        browser_service=browser_service,
        proxies=working_proxies,
    )
    export_service = ExportService(settings=settings)

    snapshot_service = SnapshotService(repository=snapshot_repository)
    comparison_service = ComparisonService()

    export_dir = str(Path(settings.export_path).parent)
    comparison_export_service = ComparisonExportService(
        export_dir=export_dir,
        events_repository=events_repository,
    )

    batch_enrichment_service = BatchEnrichmentService()
    data_cleaner_service = DataCleanerService(
        price_deviation_up=settings.price_deviation_up,
        price_deviation_down=settings.price_deviation_down,
    )

    pool_service = PoolService(pool_repository=pool_repository)
    catalog_schedule = CatalogSchedule(
        slots=list(settings.catalog_sync_times),
        state_file=Path(data_dir) / CATALOG_SYNC_STATE_FILENAME,
    )

    try:
        # --- Шаг 6: Этап 1 (парсинг каталога) — только по расписанию ---
        listings: list = []
        catalog_token: str | None = None
        catalog_executed = False

        now_utc = datetime.now(timezone.utc)

        if catalog_schedule.is_due(now_utc):
            logger.info(
                "начало_этапа_1_по_расписанию",
                step=f"слоты={catalog_schedule.describe_slots()}",
                proxies_available=len(working_proxies),
            )

            try:
                scraped_listings, scraped_token = (
                    await scraper_service.scrape_catalog()
                )

                if scraped_listings:
                    # Синхронизация пула: новые ID добавляются,
                    # метки last_seen обновляются. Удалений нет.
                    pool_service.sync_from_catalog(
                        {l.external_id for l in scraped_listings},
                    )
                    catalog_schedule.mark_synced(datetime.now(timezone.utc))

                    listings = scraped_listings
                    catalog_token = scraped_token
                    catalog_executed = True

                    logger.info(
                        "этап_1_выполнен",
                        total=len(listings),
                        token_available=catalog_token is not None,
                        step="scraping",
                    )
                else:
                    # Каталог пуст — метку расписания НЕ пишем:
                    # следующий прогон (через 2 минуты) повторит попытку.
                    logger.warning(
                        "этап_1_каталог_пуст_синхронизация_не_выполнена",
                        step="scraping",
                    )
            except Exception as e:
                logger.warning(
                    "ошибка_этапа_1_повтор_в_следующем_прогоне",
                    error=str(e)[:300],
                    error_type=type(e).__name__,
                    step="scraping",
                )
        else:
            logger.info(
                "этап_1_пропущен_не_по_расписанию",
                step=f"слоты={catalog_schedule.describe_slots()}",
            )

        # --- Шаг 6.5: Карточки из пула и БД (прогон без каталога) ---
        if not catalog_executed:
            pool_ids = pool_service.get_enrichment_ids()

            if not pool_ids:
                # Пустой пул — прогон пропускается целиком
                await browser_service.stop()
                return

            listings, missing_in_db = _load_pool_listings(pool_ids, repository)

            if missing_in_db > 0:
                logger.info(
                    "id_из_пула_без_записей_в_бд_пропущены",
                    count=missing_in_db,
                    step="появятся после ближайшей синхронизации каталога",
                )

            if not listings:
                logger.warning(
                    "нет_карточек_для_обработки_прогон_завершён",
                    pool_total=len(pool_ids),
                    step="enrichment",
                )
                await browser_service.stop()
                return

            logger.info(
                "карточки_загружены_из_бд_по_пулу",
                total=len(listings),
                pool_total=len(pool_ids),
                step="enrichment",
            )

        # --- Шаг 7: Batch-обогащение через API (Этап 2a) ---
        batch_enriched_count = 0

        if settings.use_proxy and working_proxies:
            # ── Параллельный batch: N прокси-браузеров ──
            await browser_service.stop()

            first_search_url = settings.search_urls[0]

            logger.info(
                "начало_batch_обогащения_parallel",
                total=len(listings),
                step=f"воркеров={min(len(working_proxies), settings.max_proxy_workers)}",
            )

            try:
                await batch_enrichment_service.enrich_batch_parallel(
                    settings=settings,
                    listings=listings,
                    proxies=working_proxies,
                    search_url=first_search_url,
                    proxy_service=proxy_service,
                )

                batch_enriched_count = _count_enriched(listings)
                batch_fatal_count = sum(
                    1 for l in listings
                    if l.enrichment_skip_reason is not None
                )
                batch_unenriched = _count_unenriched(listings)

                logger.info(
                    "batch_обогащение_parallel_завершено",
                    step=f"обогащено={batch_enriched_count}, "
                         f"фатальных={batch_fatal_count}, "
                         f"для_retry={batch_unenriched}",
                )

            except Exception as e:
                logger.warning(
                    "ошибка_batch_обогащения_parallel",
                    error=str(e)[:300],
                    error_type=type(e).__name__,
                    step="batch_enrichment",
                )

        elif catalog_token is not None:
            # ── Однопоточный batch: токен от каталога ──
            try:
                catalog_page_alive = await browser_service.is_alive()

                if catalog_page_alive:
                    catalog_page = browser_service.page

                    logger.info(
                        "начало_batch_обогащения",
                        total=len(listings),
                        step="batch_enrichment",
                    )

                    await batch_enrichment_service.enrich_batch(
                        page=catalog_page,
                        token=catalog_token,
                        listings=listings,
                        search_url=settings.search_urls[0],
                    )

                    batch_enriched_count = _count_enriched(listings)
                    batch_fatal_count = sum(
                        1 for l in listings
                        if l.enrichment_skip_reason is not None
                    )
                    batch_unenriched = _count_unenriched(listings)

                    logger.info(
                        "batch_обогащение_завершено",
                        step=f"обогащено={batch_enriched_count}, "
                             f"фатальных={batch_fatal_count}, "
                             f"для_retry={batch_unenriched}",
                    )
                else:
                    logger.warning(
                        "браузер_каталога_не_жив_переход_к_standalone",
                        step="batch_enrichment",
                    )

            except Exception as e:
                logger.warning(
                    "ошибка_batch_обогащения",
                    error=str(e)[:300],
                    error_type=type(e).__name__,
                    step="batch_enrichment",
                )

            # Закрываем браузер каталога
            await browser_service.stop()
        else:
            # ── Standalone batch: прогон без каталога (нет токена этапа 1) ──
            # Один браузер без прокси: загрузка страницы поиска,
            # перехват токена, batch-обогащение.
            await browser_service.stop()

            logger.info(
                "начало_batch_обогащения_standalone",
                total=len(listings),
                step="batch_enrichment",
            )

            await _run_standalone_batch(
                settings=settings,
                batch_enrichment_service=batch_enrichment_service,
                listings=listings,
                logger=logger,
                purpose="batch",
            )

            batch_enriched_count = _count_enriched(listings)
            batch_fatal_count = sum(
                1 for l in listings
                if l.enrichment_skip_reason is not None
            )
            batch_unenriched = _count_unenriched(listings)

            logger.info(
                "batch_обогащение_standalone_завершено",
                step=f"обогащено={batch_enriched_count}, "
                     f"фатальных={batch_fatal_count}, "
                     f"для_retry={batch_unenriched}",
            )

        # --- Шаг 8: Batch-retry необработанных (один раз, без прокси) ---
        unenriched_listings = _get_unenriched_listings(listings)

        if unenriched_listings:
            logger.info(
                "начало_batch_retry",
                step=f"необогащённых={len(unenriched_listings)}, "
                     f"уже_обогащено={batch_enriched_count}",
            )

            retry_ok = await _run_standalone_batch(
                settings=settings,
                batch_enrichment_service=batch_enrichment_service,
                listings=unenriched_listings,
                logger=logger,
                purpose="batch_retry",
            )

            if retry_ok:
                # Финальная статистика после retry
                final_enriched = _count_enriched(listings)
                final_unenriched = _count_unenriched(listings)
                final_fatal = sum(
                    1 for l in listings if l.enrichment_skip_reason is not None
                )

                if final_unenriched > 0:
                    logger.info(
                        "batch_retry_остались_необработанные_пропускаем",
                        step=f"обогащено={final_enriched}, "
                             f"фатальных={final_fatal}, "
                             f"пропущено={final_unenriched}",
                    )
                else:
                    logger.info(
                        "все_карточки_обработаны_после_retry",
                        step=f"обогащено={final_enriched}, "
                             f"фатальных={final_fatal}",
                    )
            else:
                logger.warning(
                    "batch_retry_не_выполнен_токен_недоступен",
                    step=f"пропущено={len(unenriched_listings)}",
                )
        else:
            logger.info(
                "все_карточки_обогащены_batch_retry_не_нужен",
                step=f"обогащено={batch_enriched_count}",
            )

        logger.info(
            "карточки_обработаны",
            total=len(listings),
            step="enrichment",
        )

        # --- Шаг 8.5: Расчёт price_per_sqm ---
        data_cleaner_service.clean_listings(listings)

        # --- Шаг 9: Сохранение в базу данных ---
        # Фильтрация: не сохраняем карточки с пустыми данными (без фатальной
        # причины), чтобы не перезаписать ранее собранные корректные данные.
        # Такие карточки появляются при ошибках API (api_false, протухший токен).
        listings_to_save: list = []
        skipped_empty_count = 0

        for listing in listings:
            # Карточки с фатальной причиной — сохраняем (чтобы skip_reason
            # зафиксировался в БД и карточка не обрабатывалась повторно).
            if listing.enrichment_skip_reason is not None:
                listings_to_save.append(listing)
                continue

            # Проверяем наличие данных: календарь ИЛИ цены должны быть непустыми.
            has_calendar = (
                listing.calendar_60_days
                and any(c != 0 for c in listing.calendar_60_days)
            )
            has_prices = (
                listing.prices_60_days
                and any(p != 0 for p in listing.prices_60_days)
            )

            if has_calendar or has_prices:
                listings_to_save.append(listing)
            else:
                skipped_empty_count += 1

        if skipped_empty_count > 0:
            logger.warning(
                "пропущены_пустые_карточки_перед_сохранением",
                step=f"пропущено={skipped_empty_count}, "
                     f"сохраняется={len(listings_to_save)}, "
                     f"всего={len(listings)}",
            )

        # snapshot_date должен отражать момент фактической записи в БД.
        # В прогонах без каталога карточки загружаются из БД со старой
        # snapshot_date и записывались обратно без обновления — из-за
        # этого дашборд показывал устаревшую дату обновления данных.
        save_moment = datetime.now(timezone.utc)
        for listing in listings_to_save:
            listing.snapshot_date = save_moment

        logger.info("сохранение_в_бд", step="storage")
        saved_count = repository.upsert_many(listings_to_save)
        logger.info(
            "данные_сохранены",
            total=saved_count,
            step="storage",
        )

        # Шаг 9.5 (удаление отсутствующих объявлений) отключён:
        # пул и БД только накапливают записи, очистка не выполняется.

        # --- Шаг 10: Сохранение снимков ---
        logger.info("сохранение_снимков", step="snapshots")
        all_listings = repository.get_all()
        snapshot_service.save_snapshots(all_listings)

        # --- Шаг 11: Сравнение снимков ---
        all_events = _run_comparison(
            listings=all_listings,
            snapshot_repository=snapshot_repository,
            comparison_service=comparison_service,
            logger=logger,
        )

        if all_events:
            logger.info(
                "экспорт_отчёта_сравнения",
                total=len(all_events),
                step="comparison_export",
            )
            comparison_path = comparison_export_service.export(all_events)
            logger.info(
                "отчёт_сравнения_сохранён",
                path=comparison_path,
                step="comparison_export",
            )
        else:
            logger.info(
                "событий_не_обнаружено_отчёт_не_создан",
                step="comparison_export",
            )

        # --- Шаг 12: Экспорт отчётов ---
        logger.info("экспорт_в_excel", step="export")

        export_path = export_service.export(all_listings)
        logger.info(
            "основной_отчёт_сохранён",
            path=export_path,
            total=len(all_listings),
            step="export",
        )

        run_timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        base_name = Path(settings.export_path).stem
        snapshot_filename = f"{base_name}_{run_timestamp}.xlsx"
        snapshot_path = str(Path(export_dir) / snapshot_filename)

        export_service.export(all_listings, output_path=snapshot_path)
        logger.info(
            "датированный_отчёт_сохранён",
            path=snapshot_path,
            total=len(all_listings),
            step="export",
        )

    finally:
        try:
            if await browser_service.is_alive():
                await browser_service.stop()
        except Exception:
            pass

        # ── Экспорт статусов прокси для админки ──
        # Файл proxies_status.json перезаписывается после каждого прогона:
        # к этому моменту часть прокси могла быть исключена из пула
        # рабочих (mark_dead) по ходу обработки. Статусы нужны только
        # админке (rentpuls.ru) и не влияют на работу парсера.
        if proxy_service is not None:
            try:
                proxy_service.export_status()
            except Exception as e:
                logger.warning(
                    "ошибка_экспорта_статусов_прокси",
                    error=str(e)[:300],
                    error_type=type(e).__name__,
                    step="cleanup",
                )

        repository.close()
        snapshot_repository.close()
        pool_repository.close()

        # Закрытие пула событий сравнения (безопасно, даже если не создан)
        if events_repository is not None:
            try:
                events_repository.close()
            except Exception as e:
                logger.warning(
                    "ошибка_закрытия_репозитория_событий",
                    error=str(e)[:300],
                    error_type=type(e).__name__,
                    step="cleanup",
                )


async def run_loop() -> None:
    """Бесконечный цикл прогонов парсера.

    Выполняет run() в цикле с паузой _CYCLE_PAUSE_SECONDS между прогонами.
    При критической ошибке — записывает статус "failed" и завершает процесс.
    При SIGTERM/SIGINT — корректно завершает текущий прогон и выходит.

    Перед каждым прогоном проверяется ночная пауза (PAUSE_START–PAUSE_END).
    Если текущее время МСК попадает в окно паузы — парсер ожидает до конца
    паузы с проверкой SIGTERM каждую секунду.

    Каждый прогон записывает результат в JSON-файл (data/last_run_status.json),
    который админка мониторит для отображения истории запусков.
    """
    global _shutdown_requested  # noqa: PLW0603

    # Регистрируем обработчики сигналов для корректной остановки
    signal.signal(signal.SIGTERM, _request_shutdown)
    signal.signal(signal.SIGINT, _request_shutdown)

    # Загружаем settings один раз для определения data_dir и паузы
    try:
        settings = Settings.load()
    except RuntimeError as e:
        print(f"[ОШИБКА] {e}", file=sys.stderr)  # noqa: T201
        sys.exit(1)

    configure_logging(
        log_level=settings.log_level,
        log_file_path=settings.log_file_path,
    )
    logger = get_logger("main")

    data_dir = str(Path(settings.export_path).parent)
    run_number = 0

    # Формируем описание слотов каталога для приветственного лога
    slots_str = ", ".join(
        f"{h:02d}:{m:02d}" for h, m in settings.catalog_sync_times
    )

    # Логируем настройки паузы
    if settings.pause_start and settings.pause_end:
        logger.info(
            "цикл_парсера_запущен",
            step=f"пауза_между_прогонами={_CYCLE_PAUSE_SECONDS}с, "
                 f"ночная_пауза={settings.pause_start[0]:02d}:{settings.pause_start[1]:02d}"
                 f"–{settings.pause_end[0]:02d}:{settings.pause_end[1]:02d} МСК, "
                 f"слоты_каталога={slots_str} МСК",
        )
    else:
        logger.info(
            "цикл_парсера_запущен",
            step=f"пауза_между_прогонами={_CYCLE_PAUSE_SECONDS}с, "
                 f"ночная_пауза=отключена, "
                 f"слоты_каталога={slots_str} МСК",
        )

    while not _shutdown_requested:
        # ── Проверка ночной паузы перед стартом прогона ──
        if settings.pause_start and settings.pause_end:
            await _wait_for_pause_end(
                pause_start=settings.pause_start,
                pause_end=settings.pause_end,
                logger=logger,
            )

            # После ожидания паузы проверяем, не запросили ли остановку
            if _shutdown_requested:
                logger.info(
                    "остановка_запрошена_после_ночной_паузы",
                    step="shutdown",
                )
                break

        run_number += 1
        started_at = datetime.now(timezone.utc)
        started_at_iso = started_at.isoformat()

        logger.info("═" * 60)
        logger.info(
            "начало_прогона",
            step=f"прогон={run_number}",
        )

        run_start_time = time.perf_counter()

        try:
            await run()

            run_elapsed = time.perf_counter() - run_start_time
            finished_at_iso = datetime.now(timezone.utc).isoformat()

            # Читаем статистику из БД для JSON-статуса
            listings_count = 0
            events_count = 0
            try:
                stats_repos = create_repositories(settings)
                listings_count = stats_repos.listing.count()
                stats_repos.listing.close()
                stats_repos.snapshot.close()
                stats_repos.pool.close()
            except Exception:
                pass

            # Считаем события из последнего отчёта сравнения
            try:
                comparison_dir = Path(data_dir)
                comparison_files = sorted(
                    comparison_dir.glob("comparison_report_*.xlsx"),
                    reverse=True,
                )
                if comparison_files:
                    from openpyxl import load_workbook

                    wb = load_workbook(
                        str(comparison_files[0]), read_only=True
                    )
                    ws = wb.active
                    events_count = max(0, ws.max_row - 1) if ws.max_row else 0
                    wb.close()
            except Exception:
                pass

            _write_run_status(
                data_dir=data_dir,
                status="success",
                started_at=started_at_iso,
                finished_at=finished_at_iso,
                listings_count=listings_count,
                events_count=events_count,
                run_number=run_number,
            )

            logger.info(
                "прогон_завершён_успешно",
                step=f"прогон={run_number}, "
                     f"время={_format_duration(run_elapsed)}, "
                     f"объявлений={listings_count}, "
                     f"событий={events_count}",
            )

        except KeyboardInterrupt:
            finished_at_iso = datetime.now(timezone.utc).isoformat()
            _write_run_status(
                data_dir=data_dir,
                status="cancelled",
                started_at=started_at_iso,
                finished_at=finished_at_iso,
                run_number=run_number,
            )
            logger.warning(
                "прогон_прерван_пользователем",
                step=f"прогон={run_number}",
            )
            break

        except Exception as e:
            run_elapsed = time.perf_counter() - run_start_time
            finished_at_iso = datetime.now(timezone.utc).isoformat()

            error_msg = f"{type(e).__name__}: {e}"

            _write_run_status(
                data_dir=data_dir,
                status="failed",
                started_at=started_at_iso,
                finished_at=finished_at_iso,
                error=error_msg[:500],
                run_number=run_number,
            )

            logger.exception(
                "критическая_ошибка_прогона",
                error=str(e),
                error_type=type(e).__name__,
                step=f"прогон={run_number}, "
                     f"время={_format_duration(run_elapsed)}",
            )
            sys.exit(1)

        # ── Проверка флага остановки перед паузой ──
        if _shutdown_requested:
            logger.info(
                "остановка_запрошена_после_прогона",
                step=f"прогон={run_number}",
            )
            break

        # ── Пауза между прогонами с проверкой SIGTERM каждую секунду ──
        logger.info(
            "пауза_между_прогонами",
            step=f"пауза={_CYCLE_PAUSE_SECONDS}с, "
                 f"следующий_прогон={run_number + 1}",
        )

        for _ in range(_CYCLE_PAUSE_SECONDS):
            if _shutdown_requested:
                logger.info(
                    "остановка_запрошена_во_время_паузы",
                    step=f"прогон={run_number}",
                )
                break
            await asyncio.sleep(1.0)

    logger.info(
        "цикл_парсера_завершён",
        step=f"прогонов_выполнено={run_number}",
    )


def _format_duration(seconds: float) -> str:
    """Форматирует длительность в человекочитаемый вид.

    Args:
        seconds: Длительность в секундах.

    Returns:
        Строка вида «Xм Yс» или «Yс».
    """
    total = int(seconds)
    minutes = total // 60
    secs = total % 60
    if minutes > 0:
        return f"{minutes}м {secs}с"
    return f"{secs}с"


def _run_comparison(
    listings: list,
    snapshot_repository: BaseSnapshotRepository,
    comparison_service: ComparisonService,
    logger: "any",  # type: ignore[name-defined]
) -> list[AnyEvent]:
    """Сравнивает последние два снимка для каждого объявления."""
    start_time = time.perf_counter()

    id_title_map: dict[str, str] = {}
    for listing in listings:
        external_id: str = getattr(listing, "external_id", "")
        title: str = getattr(listing, "title", "")
        if external_id:
            id_title_map[external_id] = title

    if not id_title_map:
        logger.info("сравнение_пропущено_нет_id", step="comparison")
        return []

    logger.info(
        "загрузка_снимков_для_сравнения",
        total=len(id_title_map),
        step="comparison",
    )

    snapshots_map = snapshot_repository.get_last_two_batch(
        list(id_title_map.keys())
    )

    load_elapsed = time.perf_counter() - start_time
    logger.info(
        "снимки_загружены",
        total_ids=len(id_title_map),
        total_with_snapshots=len(snapshots_map),
        elapsed=f"{load_elapsed:.2f}с",
        step="comparison",
    )

    all_events: list[AnyEvent] = []
    compared = 0
    skipped = 0

    for external_id, title in id_title_map.items():
        snapshots = snapshots_map.get(external_id)

        if not snapshots or len(snapshots) < 2:
            skipped += 1
            continue

        old_snapshot, new_snapshot = snapshots[0], snapshots[1]
        events = comparison_service.compare(
            old_snapshot=old_snapshot,
            new_snapshot=new_snapshot,
            listing_title=title,
        )

        all_events.extend(events)
        compared += 1

    total_elapsed = time.perf_counter() - start_time

    logger.info(
        "сравнение_завершено",
        compared=compared,
        skipped=skipped,
        total_events=len(all_events),
        elapsed=f"{total_elapsed:.2f}с",
        step="comparison",
    )

    return sorted(all_events, key=lambda e: e.checkin_date)


def main() -> None:
    """Синхронная точка входа — запускает бесконечный цикл прогонов."""
    asyncio.run(run_loop())


if __name__ == "__main__":
    main()
