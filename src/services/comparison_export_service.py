"""Сервис экспорта событий бронирования и отмен в Excel и БД."""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config.logger import get_logger
from src.models.booking_event import AnyEvent, BookingEvent, CancellationEvent, EventType
from src.repositories.base_comparison_events_repository import (
    BaseComparisonEventsRepository,
)

logger = get_logger("service.comparison_export")

# Цвета строк
_COLOR_BOOKING = "C6EFCE"       # зелёный — бронь
_COLOR_CANCELLATION = "FFCCCC"  # красный — отмена
_COLOR_HEADER = "2F4F7F"        # тёмно-синий — шапка
_COLOR_HEADER_FONT = "FFFFFF"   # белый — текст шапки

# Заголовки столбцов
_HEADERS = [
    "Тип события",
    "ID объявления",
    "Название",
    "Дата сделки",
    "Дата заезда",
    "Дата выезда",
    "Ночей",
    "Глубина (дней)",
    "Цена за ночь (руб.)",
    "Итого (руб.)",
]

# Ширина столбцов (в символах)
_COLUMN_WIDTHS = [16, 18, 40, 20, 14, 14, 9, 16, 22, 18]


class ComparisonExportService:
    """Сервис экспорта событий сравнения снимков.

    Формирует файл comparison_report_<дата>.xlsx в папке экспорта.
    Строки броней выделены зелёным, отмены — красным.

    Если передан events_repository, дополнительно записывает события
    в базу данных перед экспортом в Excel. Ошибка записи в БД не
    прерывает работу — Excel остаётся резервным каналом.
    """

    def __init__(
        self,
        export_dir: str,
        events_repository: BaseComparisonEventsRepository | None = None,
    ) -> None:
        """Инициализирует сервис.

        Args:
            export_dir: Путь к папке для сохранения Excel-файла.
            events_repository: Опциональный репозиторий для записи событий в БД.
                Если None — события пишутся только в Excel (обратная совместимость).
        """
        self._export_dir = Path(export_dir)
        self._events_repository = events_repository

    def export(self, events: list[AnyEvent]) -> str:
        """Экспортирует события в БД (если репозиторий передан) и в Excel-файл.

        Порядок операций:
            1. Запись в БД через events_repository.bulk_insert (если задан).
               Ошибка не прерывает работу — логируется и игнорируется.
            2. Запись в Excel — всегда выполняется, независимо от результата БД.

        Args:
            events: Список событий BookingEvent и CancellationEvent.

        Returns:
            Путь к созданному Excel-файлу.

        Raises:
            OSError: Если не удалось создать папку или записать Excel-файл.
        """
        # ── Шаг 1: Запись в БД (best-effort, не блокирует Excel) ──
        self._save_to_database(events)

        # ── Шаг 2: Запись в Excel (основной путь) ──
        return self._save_to_excel(events)

    def _save_to_database(self, events: list[AnyEvent]) -> None:
        """Записывает события в БД через репозиторий.

        Метод рассчитан на graceful degradation: любая ошибка БД
        (сеть, БД недоступна, конфликт схемы) логируется, но не
        прерывает прогон парсера — Excel остаётся резервным путём.

        Args:
            events: Список событий для записи.
        """
        if self._events_repository is None:
            return

        if not events:
            logger.debug(
                "запись_в_бд_пропущена_пустой_список",
                step="db_export",
            )
            return

        try:
            inserted = self._events_repository.bulk_insert(events)
            logger.info(
                "события_записаны_в_бд",
                received=len(events),
                inserted=inserted,
                duplicates=len(events) - inserted,
                step="db_export",
            )
        except Exception as e:
            # Ошибку в БД проглатываем — Excel продолжит работать
            logger.warning(
                "ошибка_записи_событий_в_бд",
                error=str(e)[:300],
                error_type=type(e).__name__,
                events_count=len(events),
                step="db_export",
            )

    def _save_to_excel(self, events: list[AnyEvent]) -> str:
        """Сохраняет события в Excel-файл.

        Args:
            events: Список событий для записи.

        Returns:
            Путь к созданному файлу.

        Raises:
            OSError: Если не удалось создать папку или записать файл.
        """
        self._export_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = self._export_dir / f"comparison_report_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "События"

        self._write_header(ws)
        self._write_rows(ws, events)
        self._apply_autofilter(ws)
        self._freeze_header(ws)

        wb.save(file_path)

        logger.info(
            "отчёт_сравнения_сохранён",
            path=str(file_path),
            total=len(events),
        )

        return str(file_path)

    def _write_header(self, ws: object) -> None:
        """Записывает строку заголовков с форматированием.

        Args:
            ws: Лист Excel.
        """
        header_fill = PatternFill(
            start_color=_COLOR_HEADER,
            end_color=_COLOR_HEADER,
            fill_type="solid",
        )
        header_font = Font(
            bold=True,
            color=_COLOR_HEADER_FONT,
        )
        center = Alignment(horizontal="center", vertical="center")

        for col_idx, (title, width) in enumerate(
            zip(_HEADERS, _COLUMN_WIDTHS), start=1
        ):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 22

    def _write_rows(self, ws: object, events: list[AnyEvent]) -> None:
        """Записывает строки событий с цветовой маркировкой.

        Args:
            ws: Лист Excel.
            events: Список событий для записи.
        """
        for row_idx, event in enumerate(events, start=2):
            row_data = self._event_to_row(event)
            fill_color = (
                _COLOR_BOOKING
                if event.event_type == EventType.BOOKING
                else _COLOR_CANCELLATION
            )
            fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type="solid",
            )
            center = Alignment(horizontal="center", vertical="center")

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                # Название выравниваем по левому краю
                if col_idx == 3:
                    cell.alignment = Alignment(
                        horizontal="left", vertical="center"
                    )
                else:
                    cell.alignment = center

    def _event_to_row(self, event: AnyEvent) -> list:
        """Преобразует событие в список значений для строки Excel.

        Args:
            event: Событие бронирования или отмены.

        Returns:
            Список значений в порядке столбцов _HEADERS.
        """
        return [
            event.event_type.value,                          # Тип события
            event.listing_external_id,                       # ID объявления
            event.listing_title,                             # Название
            event.snapshot_dt.strftime("%d.%m.%Y %H:%M"),   # Дата сделки
            event.checkin_date.strftime("%d.%m.%Y"),         # Дата заезда
            event.checkout_date.strftime("%d.%m.%Y"),        # Дата выезда
            event.nights,                                    # Ночей
            event.depth_days,                                # Глубина (дней)
            event.price_per_night,                           # Цена за ночь
            event.total_price,                               # Итого
        ]

    def _apply_autofilter(self, ws: object) -> None:
        """Устанавливает автофильтр на все столбцы.

        Args:
            ws: Лист Excel.
        """
        last_col = get_column_letter(len(_HEADERS))
        ws.auto_filter.ref = f"A1:{last_col}1"

    def _freeze_header(self, ws: object) -> None:
        """Фиксирует первую строку при прокрутке.

        Args:
            ws: Лист Excel.
        """
        ws.freeze_panes = "A2"
