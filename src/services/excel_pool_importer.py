"""Импортёр ID объявлений из Excel-отчёта парсера.

Одноразовый инструмент первичного наполнения пула: читает основной
или датированный отчёт (sutochno_report*.xlsx), находит столбец
«ID объявления» и возвращает множество ID в виде строк.

Не зависит от настроек приложения — путь к файлу передаётся
в метод extract_ids() явно.
"""

from pathlib import Path

from openpyxl import load_workbook

from src.config.logger import get_logger

logger = get_logger("excel_pool_importer")

# Название столбца с ID в Excel-отчёте (см. ExportService._COLUMNS)
_ID_COLUMN_HEADER: str = "ID объявления"


class ExcelPoolImporter:
    """Извлекает ID объявлений из столбца «ID объявления» Excel-отчёта."""

    def extract_ids(self, excel_path: str | Path) -> set[str]:
        """Читает Excel-файл и возвращает множество ID объявлений.

        Ожидает формат, который формирует ExportService: первая строка —
        заголовки, среди которых есть «ID объявления». Данные начинаются
        со второй строки.

        Ячейки столбца конвертируются в строки: отчёт хранит ID как текст,
        но если файл пересохранялся в Excel, числовые ячейки приходят
        как int/float — оба варианта обрабатываются.

        Args:
            excel_path: Путь к Excel-файлу отчёта.

        Returns:
            Множество ID (уникальные непустые значения столбца).

        Raises:
            FileNotFoundError: Если файл не существует.
            RuntimeError: Если файл пуст, не открывается или столбец
                «ID объявления» не найден в заголовках.
        """
        path = Path(excel_path)

        if not path.exists():
            raise FileNotFoundError(
                f"Excel-файл не найден: {path}. "
                f"Проверьте путь к отчёту."
            )

        logger.info(
            "импорт_excel_начат",
            path=str(path),
        )

        try:
            workbook = load_workbook(str(path), read_only=True)
        except Exception as e:
            raise RuntimeError(
                f"Не удалось открыть Excel-файл {path}: {e}. "
                f"Убедитесь, что это корректный .xlsx без повреждений "
                f"и не открыт в Excel во время импорта."
            ) from e

        try:
            worksheet = workbook.active
            if worksheet is None:
                raise RuntimeError(
                    f"В Excel-файле {path} нет активного листа."
                )

            rows = worksheet.iter_rows(values_only=True)

            # ── Поиск столбца по заголовку первой строки ──
            try:
                header_row = next(rows)
            except StopIteration:
                raise RuntimeError(
                    f"Excel-файл {path} пуст: нет строки заголовков."
                ) from None

            id_column_idx = self._find_id_column(header_row)

            ids: set[str] = set()
            total_rows = 0
            skipped_cells = 0
            duplicates = 0

            for row in rows:
                total_rows += 1

                if id_column_idx >= len(row):
                    skipped_cells += 1
                    continue

                external_id = self._normalize_cell(row[id_column_idx])

                if external_id is None:
                    skipped_cells += 1
                    continue

                if external_id in ids:
                    duplicates += 1
                    continue

                ids.add(external_id)

            logger.info(
                "импорт_excel_завершён",
                path=str(path),
                total_rows=total_rows,
                unique_ids=len(ids),
                duplicates_in_file=duplicates,
                skipped_cells=skipped_cells,
            )

            return ids

        finally:
            workbook.close()

    @staticmethod
    def _find_id_column(header_row: tuple) -> int:
        """Находит индекс столбца «ID объявления» в строке заголовков.

        Args:
            header_row: Кортеж значений первой строки листа.

        Returns:
            Индекс столбца (0-based).

        Raises:
            RuntimeError: Если столбец не найден — в сообщении перечислены
                фактические заголовки для диагностики.
        """
        for idx, cell_value in enumerate(header_row):
            if isinstance(cell_value, str) and cell_value.strip() == _ID_COLUMN_HEADER:
                return idx

        actual_headers = [
            str(v).strip() if v is not None else ""
            for v in header_row
        ]
        raise RuntimeError(
            f"Столбец «{_ID_COLUMN_HEADER}» не найден в первой строке "
            f"Excel-файла. Фактические заголовки: {actual_headers[:10]}..."
        )

    @staticmethod
    def _normalize_cell(cell_value: object) -> str | None:
        """Конвертирует значение ячейки в строку ID.

        Обрабатывает варианты, которые может содержать пересохранённый
        Excel-файл:
        - str — штатный формат отчёта парсера;
        - int — Excel преобразовал текстовую ячейку в число;
        - float с целым значением — то же самое (709383.0);
        - None/прочее — некорректная ячейка.

        Args:
            cell_value: Значение ячейки из openpyxl.

        Returns:
            ID строкой или None, если ячейка не содержит корректный ID.
        """
        if isinstance(cell_value, str):
            normalized = cell_value.strip()
            return normalized if normalized else None

        if isinstance(cell_value, int):
            return str(cell_value)

        if isinstance(cell_value, float):
            if cell_value.is_integer():
                return str(int(cell_value))
            return None

        return None
