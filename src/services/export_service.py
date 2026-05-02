"""Сервис экспорта данных — формирование Excel-отчёта."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.config.logger import get_logger
from src.config.settings import Settings
from src.models.listing import RawListing

logger = get_logger("export")

# Определение столбцов отчёта
_COLUMNS: list[dict[str, str | int]] = [
    {"header": "ID объявления", "width": 14},
    {"header": "Название", "width": 50},
    {"header": "Цена (руб./сут.)", "width": 16},
    {"header": "Рейтинг", "width": 10},
    {"header": "Отзывы", "width": 10},
    {"header": "Площадь (м²)", "width": 13},
    {"header": "Гостей", "width": 9},
    {"header": "Адрес", "width": 45},
    {"header": "Метро", "width": 30},
    {"header": "Быстрое бронирование", "width": 22},
    {"header": "Занятость (%)", "width": 14},
    {"header": "Календарь 60 дней", "width": 65},
    {"header": "Средняя цена (руб./сут.)", "width": 22},
    {"header": "Цены 60 дней (руб./сут.)", "width": 65},
    {"header": "Ссылка", "width": 20},
    {"header": "Дата снимка", "width": 20},
]


class ExportService:
    """Сервис экспорта объявлений в Excel-файл.

    Создаёт отформатированную таблицу с автофильтрами,
    кликабельными ссылками и стилизованным заголовком.
    """

    def __init__(self, settings: Settings) -> None:
        """Инициализирует сервис экспорта.

        Args:
            settings: Настройки приложения (путь к выходному файлу).
        """
        self._settings = settings

    def export(self, listings: list[RawListing]) -> str:
        """Экспортирует список объявлений в Excel-файл.

        Args:
            listings: Список объявлений для экспорта.

        Returns:
            Путь к созданному файлу.

        Raises:
            RuntimeError: Если список объявлений пуст.
        """
        if not listings:
            raise RuntimeError(
                "Нет данных для экспорта. Парсинг не вернул ни одного объявления."
            )

        export_path = Path(self._settings.export_path)
        export_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = "Объявления"

        # Формируем заголовок
        self._write_header(ws)

        # Заполняем данные
        self._write_data(ws, listings)

        # Настраиваем ширину столбцов
        self._set_column_widths(ws)

        # Добавляем автофильтр
        self._add_autofilter(ws, len(listings))

        # Закрепляем первую строку
        ws.freeze_panes = "A2"

        # Сохраняем файл
        wb.save(str(export_path))

        logger.info(
            "отчёт_сохранён",
            path=str(export_path),
            total=len(listings),
        )
        return str(export_path)

    def _write_header(self, ws: Worksheet) -> None:
        """Записывает и стилизует строку заголовка.

        Args:
            ws: Рабочий лист Excel.
        """
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="2F5496",
            end_color="2F5496",
            fill_type="solid",
        )
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, column_def in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = column_def["header"]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _write_data(self, ws: Worksheet, listings: list[RawListing]) -> None:
        """Записывает данные объявлений в таблицу.

        Args:
            ws: Рабочий лист Excel.
            listings: Список объявлений.
        """
        link_font = Font(color="0563C1", underline="single")

        for row_idx, listing in enumerate(listings, start=2):
            ws.cell(row=row_idx, column=1, value=listing.external_id)
            ws.cell(row=row_idx, column=2, value=listing.title)
            ws.cell(row=row_idx, column=3, value=listing.price_per_night)
            ws.cell(row=row_idx, column=4, value=listing.rating)
            ws.cell(row=row_idx, column=5, value=listing.review_count)
            ws.cell(row=row_idx, column=6, value=listing.area_m2)
            ws.cell(row=row_idx, column=7, value=listing.guests)
            ws.cell(row=row_idx, column=8, value=listing.address)
            ws.cell(row=row_idx, column=9, value=listing.metro_station)
            ws.cell(
                row=row_idx,
                column=10,
                value="Да" if listing.has_instant_booking else "Нет",
            )

            # Занятость (%)
            ws.cell(row=row_idx, column=11, value=listing.occupancy_percent)

            # Календарь 60 дней — компактная строка
            calendar_str = "".join(str(d) for d in listing.calendar_60_days) if listing.calendar_60_days else ""
            ws.cell(row=row_idx, column=12, value=calendar_str)

            # Средняя цена (руб./сут.)
            ws.cell(row=row_idx, column=13, value=listing.average_price)

            # Цены 60 дней — через точку с запятой
            prices_str = ";".join(str(p) for p in listing.prices_60_days) if listing.prices_60_days else ""
            ws.cell(row=row_idx, column=14, value=prices_str)

            # Кликабельная ссылка
            link_cell = ws.cell(row=row_idx, column=15, value="Открыть")
            link_cell.hyperlink = listing.url
            link_cell.font = link_font

            # Дата снимка
            ws.cell(
                row=row_idx,
                column=16,
                value=listing.snapshot_date.strftime("%Y-%m-%d %H:%M"),
            )

    def _set_column_widths(self, ws: Worksheet) -> None:
        """Устанавливает ширину столбцов.

        Args:
            ws: Рабочий лист Excel.
        """
        for col_idx, column_def in enumerate(_COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = column_def["width"]

    def _add_autofilter(self, ws: Worksheet, row_count: int) -> None:
        """Добавляет автофильтр на таблицу.

        Args:
            ws: Рабочий лист Excel.
            row_count: Количество строк данных.
        """
        last_col_letter = get_column_letter(len(_COLUMNS))
        ws.auto_filter.ref = f"A1:{last_col_letter}{row_count + 1}"
