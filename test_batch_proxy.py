"""Тестовый скрипт — воспроизведение этапа 2 (batch-обогащение через прокси).

Читает первые 5000 ID объявлений из Excel-отчёта, загружает и проверяет
все прокси из proxies.txt, запускает enrich_batch_parallel() с подробной
диагностикой. Позволяет выявить проблемы с прокси и воркерами до запуска
полного прогона.

Запуск:
    cd C:\\game\\test\\my-poisk\\sutochno
    python test_batch_proxy.py
"""

import asyncio
import sys
import time
from pathlib import Path

# ── Добавляем корень проекта в sys.path ──
PROJECT_ROOT = Path(__file__).resolve().parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import load_workbook

from src.config.logger import configure as configure_logging
from src.config.logger import get_logger
from src.config.settings import Settings
from src.models.listing import RawListing
from src.models.proxy import ProxyConfig
from src.services.listing.batch_enrichment_service import BatchEnrichmentService
from src.services.proxy_service import ProxyService


# ── Константы тестового скрипта ──────────────────────────────

# Путь к Excel-файлу с ID объявлений
_EXCEL_PATH: str = r"C:\game\test\my-poisk\sutochno\data\sutochno_report_20260831_231803.xlsx"

# Путь к файлу прокси
_PROXIES_PATH: str = r"C:\game\test\my-poisk\sutochno\data\proxies.txt"

# Количество ID для тестирования
_MAX_IDS: int = 5000

# Имя столбца с ID объявлений в Excel
_ID_COLUMN_NAME: str = "ID объявления"


def _load_ids_from_excel(path: str, max_ids: int) -> list[str]:
    """Загружает ID объявлений из Excel-файла.

    Ищет столбец по заголовку в первой строке, затем читает значения
    из этого столбца (пропуская пустые ячейки).

    Args:
        path: Путь к Excel-файлу.
        max_ids: Максимальное количество ID для загрузки.

    Returns:
        Список строковых ID объявлений.

    Raises:
        RuntimeError: Если файл не найден, столбец не найден или нет данных.
    """
    excel_path = Path(path)
    if not excel_path.exists():
        raise RuntimeError(f"Excel-файл не найден: {path}")

    wb = load_workbook(str(excel_path), read_only=True, data_only=True)
    ws = wb.active

    if ws is None:
        wb.close()
        raise RuntimeError(f"В файле нет активного листа: {path}")

    # Ищем столбец с ID по заголовку в первой строке
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        wb.close()
        raise RuntimeError(f"Файл пуст — нет заголовков: {path}")

    id_col_idx: int | None = None
    for idx, cell_value in enumerate(header_row):
        if cell_value is not None and str(cell_value).strip() == _ID_COLUMN_NAME:
            id_col_idx = idx
            break

    if id_col_idx is None:
        wb.close()
        raise RuntimeError(
            f"Столбец '{_ID_COLUMN_NAME}' не найден в файле: {path}. "
            f"Заголовки: {[str(h).strip() for h in header_row if h is not None]}"
        )

    # Читаем значения из найденного столбца (начиная со второй строки)
    ids: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(ids) >= max_ids:
            break

        if id_col_idx < len(row):
            cell_value = row[id_col_idx]
            if cell_value is not None:
                str_value = str(cell_value).strip()
                if str_value:
                    # Убираем .0 если число было прочитано как float
                    if str_value.endswith(".0"):
                        str_value = str_value[:-2]
                    ids.append(str_value)

    wb.close()

    if not ids:
        raise RuntimeError(
            f"Столбец '{_ID_COLUMN_NAME}' не содержит данных: {path}"
        )

    return ids


def _create_fake_listings(ids: list[str]) -> list[RawListing]:
    """Создаёт минимальные объекты RawListing из списка ID.

    Для batch-обогащения нужны только external_id, title и url.
    Остальные поля будут заполнены в процессе обогащения.

    Args:
        ids: Список строковых ID объявлений.

    Returns:
        Список объектов RawListing.
    """
    listings: list[RawListing] = []
    for ext_id in ids:
        listing = RawListing(
            external_id=ext_id,
            title=f"Тестовое объявление {ext_id}",
            url=f"https://sutochno.ru/{ext_id}",
        )
        listings.append(listing)
    return listings


def _format_duration(seconds: float) -> str:
    """Форматирует длительность в человекочитаемый вид.

    Args:
        seconds: Длительность в секундах.

    Returns:
        Строка вида «Xм Yс» или «Yс».
    """
    total = int(seconds)
    minutes = total // 60
    secs = total % 60
    if minutes > 0:
        return f"{minutes}м {secs}с"
    return f"{secs}с"


async def main() -> None:
    """Основная логика тестового скрипта."""
    # ── Шаг 1: Загрузка конфигурации ──
    settings = Settings.load()

    configure_logging(
        log_level="DEBUG",
        log_file_path=settings.log_file_path,
    )
    logger = get_logger("test_batch")

    logger.info("=" * 60)
    logger.info(
        "тестовый_скрипт_начат",
        step="batch-обогащение через прокси",
    )

    # ── Шаг 2: Загрузка ID из Excel ──
    logger.info(
        "загрузка_id_из_excel",
        step=f"файл={_EXCEL_PATH}, лимит={_MAX_IDS}",
    )

    try:
        ids = _load_ids_from_excel(_EXCEL_PATH, _MAX_IDS)
    except RuntimeError as e:
        logger.error("ошибка_загрузки_excel", error=str(e))
        return

    logger.info(
        "id_загружены",
        step=f"всего={len(ids)}, первые_5={ids[:5]}, последние_5={ids[-5:]}",
    )

    # ── Шаг 3: Создание RawListing из ID ──
    listings = _create_fake_listings(ids)
    logger.info(
        "карточки_созданы",
        step=f"всего={len(listings)}",
    )

    # ── Шаг 4: Загрузка и проверка прокси ──
    logger.info(
        "загрузка_прокси",
        step=f"файл={_PROXIES_PATH}",
    )

    proxy_service = ProxyService(settings=settings)

    try:
        proxies = proxy_service.load_proxies()
    except RuntimeError as e:
        logger.error("ошибка_загрузки_прокси", error=str(e))
        return

    logger.info(
        "прокси_загружены_начинаем_проверку",
        step=f"всего={len(proxies)}",
    )

    check_start = time.perf_counter()
    working_proxies = await proxy_service.check_proxies(proxies)
    check_elapsed = time.perf_counter() - check_start

    logger.info(
        "проверка_прокси_завершена",
        step=f"рабочих={len(working_proxies)} из {len(proxies)}, "
             f"время={_format_duration(check_elapsed)}",
    )

    if not working_proxies:
        logger.error("нет_рабочих_прокси_тест_невозможен")
        return

    # ── Диагностика: сколько воркеров будет запущено ──
    max_workers = min(len(working_proxies), settings.max_proxy_workers)

    logger.info("=" * 60)
    logger.info(
        "ДИАГНОСТИКА_ПЕРЕД_ЗАПУСКОМ",
        step=f"рабочих_прокси={len(working_proxies)}, "
             f"max_proxy_workers={settings.max_proxy_workers}, "
             f"итого_воркеров={max_workers}, "
             f"свободных_прокси_для_замены={len(working_proxies) - max_workers}",
    )
    logger.info("=" * 60)

    if len(working_proxies) > max_workers:
        logger.info(
            "ВНИМАНИЕ_прокси_отсечены",
            step=f"из {len(working_proxies)} рабочих прокси будут использованы "
                 f"только {max_workers} (ограничение MAX_PROXY_WORKERS). "
                 f"Оставшиеся {len(working_proxies) - max_workers} — резерв для замены.",
        )

    # ── Шаг 5: Запуск batch-обогащения ──
    first_search_url = settings.search_urls[0]

    logger.info(
        "начало_batch_обогащения_parallel",
        step=f"карточек={len(listings)}, воркеров={max_workers}, "
             f"search_url={first_search_url[:80]}...",
    )

    batch_service = BatchEnrichmentService()
    enrich_start = time.perf_counter()

    try:
        await batch_service.enrich_batch_parallel(
            settings=settings,
            listings=listings,
            proxies=working_proxies,
            search_url=first_search_url,
            proxy_service=proxy_service,
        )
    except Exception as e:
        logger.error(
            "ошибка_batch_обогащения",
            error=str(e)[:500],
            error_type=type(e).__name__,
        )

    enrich_elapsed = time.perf_counter() - enrich_start

    # ── Шаг 6: Подробная статистика ──
    logger.info("=" * 60)
    logger.info("ИТОГОВАЯ_СТАТИСТИКА")
    logger.info("=" * 60)

    # Подсчёт результатов
    enriched_count = 0
    fatal_count = 0
    empty_count = 0
    has_calendar_count = 0
    has_prices_count = 0
    busy_count = 0
    unbusy_count = 0

    skip_reasons: dict[str, int] = {}

    for listing in listings:
        if listing.enrichment_skip_reason is not None:
            fatal_count += 1
            reason = listing.enrichment_skip_reason
            skip_reasons[reason] = skip_reasons.get(reason, 0) + 1
            continue

        has_calendar = (
            listing.calendar_60_days
            and any(c == 1 for c in listing.calendar_60_days)
        )
        has_prices = (
            listing.prices_60_days
            and any(p > 0 for p in listing.prices_60_days)
        )

        if has_calendar or has_prices:
            enriched_count += 1
            if has_calendar:
                has_calendar_count += 1
                busy_days = sum(listing.calendar_60_days)
                if busy_days > 0:
                    busy_count += 1
            if has_prices:
                has_prices_count += 1
        else:
            empty_count += 1

        # Считаем unbusy (календарь весь нулевой, но есть цены)
        if (
            listing.calendar_60_days
            and all(c == 0 for c in listing.calendar_60_days)
            and has_prices
        ):
            unbusy_count += 1

    total_time = check_elapsed + enrich_elapsed

    logger.info(
        "результат_обогащения",
        step=f"обогащено={enriched_count}/{len(listings)} "
             f"({enriched_count * 100 / len(listings):.1f}%)",
    )
    logger.info(
        "детали_обогащения",
        step=f"с_календарём={has_calendar_count}, "
             f"с_ценами={has_prices_count}, "
             f"busy={busy_count}, "
             f"unbusy={unbusy_count}",
    )
    logger.info(
        "фатальные_ошибки",
        step=f"всего={fatal_count}, причины={skip_reasons}",
    )
    logger.info(
        "пустые_карточки",
        step=f"необогащённые={empty_count}",
    )
    logger.info(
        "время_выполнения",
        step=f"проверка_прокси={_format_duration(check_elapsed)}, "
             f"обогащение={_format_duration(enrich_elapsed)}, "
             f"общее={_format_duration(total_time)}",
    )
    logger.info(
        "прокси_статистика",
        step=f"загружено={len(proxies)}, "
             f"рабочих={len(working_proxies)}, "
             f"воркеров={max_workers}, "
             f"осталось_в_пуле={len(proxy_service.working_proxies)}",
    )

    # ── Вывод в консоль для удобства ──
    print("\n" + "=" * 60)  # noqa: T201
    print("ИТОГОВАЯ СТАТИСТИКА ТЕСТОВОГО ПРОГОНА")  # noqa: T201
    print("=" * 60)  # noqa: T201
    print(f"ID из Excel:              {len(ids)}")  # noqa: T201
    print(f"Прокси загружено:         {len(proxies)}")  # noqa: T201
    print(f"Прокси рабочих:           {len(working_proxies)}")  # noqa: T201
    print(f"MAX_PROXY_WORKERS:        {settings.max_proxy_workers}")  # noqa: T201
    print(f"Воркеров запущено:        {max_workers}")  # noqa: T201
    print(f"Прокси для замены:        {len(working_proxies) - max_workers}")  # noqa: T201
    print(f"Прокси в пуле после:      {len(proxy_service.working_proxies)}")  # noqa: T201
    print("-" * 60)  # noqa: T201
    print(f"Обогащено:                {enriched_count}/{len(listings)} "  # noqa: T201
          f"({enriched_count * 100 / len(listings):.1f}%)")
    print(f"  с календарём:           {has_calendar_count}")  # noqa: T201
    print(f"  с ценами:               {has_prices_count}")  # noqa: T201
    print(f"  busy (занятые дни):     {busy_count}")  # noqa: T201
    print(f"  unbusy (все свободны):  {unbusy_count}")  # noqa: T201
    print(f"Фатальных ошибок:         {fatal_count}")  # noqa: T201
    if skip_reasons:
        for reason, count in sorted(skip_reasons.items()):
            print(f"  {reason}: {count}")  # noqa: T201
    print(f"Необогащённых:            {empty_count}")  # noqa: T201
    print("-" * 60)  # noqa: T201
    print(f"Время проверки прокси:    {_format_duration(check_elapsed)}")  # noqa: T201
    print(f"Время обогащения:         {_format_duration(enrich_elapsed)}")  # noqa: T201
    print(f"Общее время:              {_format_duration(total_time)}")  # noqa: T201
    print("=" * 60)  # noqa: T201


if __name__ == "__main__":
    asyncio.run(main())
